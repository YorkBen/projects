# pip install openpyxl
from openpyxl import load_workbook
import json

# 加载json数据
json_data = ''
with open(r'汇总结果_231.json') as f:
    json_data = json.load(f, strict=False)

def get_max_num(json_data):
    max_len, max_len2 = 0, 0
    for item in json_data:
        if '日常病程' in item and len(item['日常病程']) > max_len:
            max_len = len(item['日常病程'])
        if '实验室数据' in item and len(item['实验室数据']) > max_len2:
            max_len2 = len(item['实验室数据'])
    return max_len, max_len2

def get_json_value(item, key):
    if key not in item:
        return ''
    else:
        return str(item[key])

columns = [
    '医保编号',
    ['性别', '年龄', '入院时间', '主诉', '现病史', '既往史', '手术外伤史', '输血史', '药物过敏史',
     '个人史', '婚育史', '月经史', '家族史', '体格检查', '专科情况（体检）', '病史小结'],
    'US',
    ['CT', 'MR', 'CR/DR、移动DR', 'X线'],
    '病理',
    ['中性粒细胞%', '白细胞', '超敏C-反应蛋白', '降钙素原', '脂肪酶', '淀粉酶', 'β-绒毛膜促性腺激素',
     '总胆红素', '天冬氨酸氨基转移酶', '丙氨酸氨基转移酶', '碱性磷酸酶', 'γ-谷氨酰转移酶', '血沉', '红细胞']
]

# 写excel
workbook = load_workbook(r"../result.xlsx")
sheet = workbook["Sheet1"]

# 写表头
ind = 1
for col in columns:
    if isinstance(col, list):
        for col2 in col:
            sheet.cell(1, ind).value = col2
            ind = ind + 1
    else:
        sheet.cell(1, ind).value = col
        ind = ind + 1

## 写内容 #############
for ind, item in enumerate(json_data):
    rn = ind + 2
    # 医保编号
    sheet.cell(rn, 1).value = get_json_value(item, '医保编号')

    # 入院记录
    cn = 2
    for col in columns[1]:
        sheet.cell(rn, cn).value = get_json_value(item['入院记录'], col)
        cn = cn + 1

    # 超声
    str_cs = ''
    if '超声' in item:
        for item_cs in item['超声']:
            for item_cs_sj in item_cs['数据']:
                arr = item_cs_sj.split(',')
                if len(arr) == 3:
                    arr.append('')
                str_cs = str_cs + '检查部位：%s\n影像描述：%s\n影像结论：%s\n\n' % (arr[1], arr[2], arr[3])
    sheet.cell(rn, cn).value = str_cs
    cn = cn + 1

    # 放射
    str_ct, str_mr, str_dr, str_xl = '', '', '', ''
    if '放射' in item:
        for item_fs in item['放射']:
            for item_fs_sj in item_fs['数据']:
                arr = item_fs_sj.split(',')
                if len(arr) == 3:
                    arr.append('')
                type = arr[0].upper()
                if 'CT' in type:
                    str_ct = str_ct + '检查名称：%s\n影像表现：%s\n影像诊断：%s\n\n' % (arr[1], arr[2], arr[3])
                elif 'MR' in type:
                    str_mr = str_mr + '检查名称：%s\n影像表现：%s\n影像诊断：%s\n\n' % (arr[1], arr[2], arr[3])
                elif 'DR' in type or 'CR' in type:
                    str_dr = str_dr + '检查名称：%s\n影像表现：%s\n影像诊断：%s\n\n' % (arr[1], arr[2], arr[3])
                elif 'X线' in type:
                    str_xl = str_xl + '检查名称：%s\n影像表现：%s\n影像诊断：%s\n\n' % (arr[1], arr[2], arr[3])
    sheet.cell(rn, cn).value = str_ct
    sheet.cell(rn, cn+1).value = str_mr
    sheet.cell(rn, cn+2).value = str_dr
    sheet.cell(rn, cn+3).value = str_xl
    cn = cn + 4

    # 病理
    str_bl = ''
    if '病理' in item:
        for item_bl in item['病理']:
            for item_bl_sj in item_bl['数据']:
                arr = item_bl_sj.split(',')
                str_bl = str_bl + '检查所见：%s\n诊断意见：%s\n\n' % (arr[0], arr[1])
    sheet.cell(rn, cn).value = str_bl
    cn = cn + 1

    # 检验
    str_dict, cols_jy = {}, columns[5]
    cols_jy_set = set(cols_jy)
    if '检验' in item:
        for item_jy in item['检验']:
            for item_jy_sj in item_jy['数据']:
                arr = item_jy_sj.split(',')
                key = arr[2]
                if key in cols_jy_set:
                    if key not in str_dict:
                        str_dict[key] = ''
                    str_dict[key] = str_dict[key] + ','.join(arr[-3:]) + '\n'

    for col in cols_jy:
        if col in str_dict:
            sheet.cell(rn, cn).value = str_dict[col]
        cn = cn + 1


workbook.save(r"r.xlsx")
