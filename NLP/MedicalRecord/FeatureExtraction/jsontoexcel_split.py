# pip install openpyxl
from openpyxl import load_workbook
import json
import re
import os
import argparse

def get_max_num(json_data):
    max_len, max_len2 = 0, 0
    for item in json_data:
        if '日常病程' in item and len(item['日常病程']) > max_len:
            max_len = len(item['日常病程'])
        if '实验室数据' in item and len(item['实验室数据']) > max_len2:
            max_len2 = len(item['实验室数据'])
    return max_len, max_len2

def get_json_value(item, key):
    if key not in item:
        return ''
    else:
        return str(item[key])

def check_cs_part(part_str):
    """
    根据超声检查部位，返回子列号
    """
    if '腹股沟' in part_str:
        return 0
    elif re.search('((肾)|(输尿)|(膀胱)|(泌尿))', part_str):
        return 1
    elif re.search('((心)|(室壁))', part_str):
        return 2
    elif re.search('(前列腺)', part_str):
        return 3
    elif re.search('[肝胆胰脾]', part_str):
        return 4
    elif re.search('(胸水)', part_str):
        return 5
    elif re.search('(腹水)', part_str):
        return 6
    elif re.search('(腹部包块)', part_str):
        return 7
    elif re.search('(腹部大血管)', part_str):
        return 8
    elif re.search('((阑尾)|(肠管))', part_str):
        return 9
    else:
        return 10

def check_ct_parts(part_str):
    """
    根据CT部位字符串，返回子列号，可能有多个
    """
    sub_inds = []
    if '胸' in part_str:
        sub_inds.append(0)
    if '腹' in part_str:
        sub_inds.append(1)
    if '上腹' in part_str:
        sub_inds.append(2)
    if '下腹' in part_str:
        sub_inds.append(3)
    if re.search('((肾)|(输尿)|(膀胱)|(泌尿)|(CTU))', part_str):
        sub_inds.append(4)
    if '盆腔' in part_str:
        sub_inds.append(5)
    if '颅脑' in part_str:
        sub_inds.append(6)

    if len(sub_inds) == 0:
        sub_inds.append(7)

    return sub_inds


def check_mr_parts(part_str):
    """
    根据MR部位字符串，返回子列号，可能有多个
    """
    sub_inds = []
    if '肝' in part_str:
        sub_inds.append(0)
    if '胆' in part_str:
        sub_inds.append(1)
    if '胰' in part_str:
        sub_inds.append(2)
    if '脾' in part_str:
        sub_inds.append(3)
    if '上腹' in part_str:
        sub_inds.append(4)
    if '下腹' in part_str:
        sub_inds.append(5)
    if 'MRCP' in part_str.upper():
        sub_inds.append(6)
    if '颅脑' in part_str:
        sub_inds.append(7)
    if '盆腔' in part_str:
        sub_inds.append(8)

    if len(sub_inds) == 0:
        sub_inds.append(9)

    return sub_inds


def check_dr_parts(part_str):
    """
    根据DR部位字符串，返回子列号，可能有多个
    """
    sub_inds = []
    if '胸' in part_str:
        sub_inds.append(0)
    if '腹' in part_str:
        sub_inds.append(1)

    if len(sub_inds) == 0:
        sub_inds.append(2)

    return sub_inds


def check_xx_parts(part_str):
    """
    根据X线部位字符串，返回子列号，可能有多个
    """
    sub_inds = [0]

    return sub_inds

def check_cell_empty(value):
    if value is None:
        return True
    elif value == 'None':
        return True
    elif value == '':
        return True
    return False


columns = [
    '医保编号',
    ['患者姓名', '性别', '年龄', '入院时间', '主诉', '现病史', '既往史', '手术外伤史', '输血史', '药物过敏史', \
     '个人史', '婚育史', '月经史', '家族史', '体格检查', '专科情况（体检）', '病史小结'],
    '出院诊断',
    ['超声_腹股沟', '超声_泌尿', '超声_心脏', '超声_前列腺', '超声_肝胆', '超声_胸水', '超声_腹水', '超声_腹部包块', '超声_腹部大血管', '超声_阑尾及肠管探查', '超声_其它'],
    ['CT_胸部', 'CT_全腹', 'CT_上腹', 'CT_下腹', 'CT_泌尿', 'CT_盆腔', 'CT_颅脑', 'CT_其它'],
    ['MR_肝', 'MR_胆', 'MR_胰', 'MR_脾', 'MR_上腹', 'MR_下腹', 'MR_MRCP', 'MR_颅脑', 'MR_盆腔', 'MR_其它'],
    ['DR_胸部', 'DR_腹部', 'DR_其它'],
    ['X线'],
    '病理',
    ['中性粒细胞%', '白细胞', '超敏C-反应蛋白', '降钙素原', '脂肪酶', '淀粉酶', 'β-绒毛膜促性腺激素', \
     '总胆红素', '天冬氨酸氨基转移酶', '丙氨酸氨基转移酶', '碱性磷酸酶', 'γ-谷氨酰转移酶', '血沉', '红细胞']
     # ,'辅助检查_外院'
]


if __name__ == '__main__':
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='2409', help='postfix num')
    parser.add_argument('-t', type=str, default='腹痛', help='数据类型')
    args = parser.parse_args()

    postfix = args.p
    data_type = args.t
    if not os.path.exists('data/%s' % data_type):
        print('data type: %s not exists' % data_type)
        exit()
    print("postfix: %s, data_type: %s" % (data_type, postfix))
    if not os.path.exists('data/%s/labeled_ind_%s.txt' % (data_type, postfix)):
        print('mrnos file: data/%s/labeled_ind_%s.txt not exists!' % (data_type, postfix))
        exit()

    ## 加载json数据 ############
    json_data = ''
    with open(r'data/%s/汇总结果_%s.json' % (data_type, postfix)) as f:
        json_data = json.load(f, strict=False)

    ## 写excel ############
    workbook = load_workbook(r"data/result.xlsx")
    sheet = workbook["Sheet1"]

    ## 写表头 #############
    ind = 1
    for col in columns:
        if isinstance(col, list):
            for col2 in col:
                sheet.cell(1, ind).value = col2
                ind = ind + 1
        else:
            sheet.cell(1, ind).value = col
            ind = ind + 1

    ## 写内容 #############
    for ind, item in enumerate(json_data):
        rn = ind + 2
        # 医保编号
        sheet.cell(rn, 1).value = get_json_value(item, '医保编号')

        # 入院记录
        cn = 2
        for col in columns[1]:
            if '入院记录' in item:
                if col == '患者姓名':
                    sheet.cell(rn, cn).value = get_json_value(item['入院记录']['病史小结'], col)
                else:
                    sheet.cell(rn, cn).value = get_json_value(item['入院记录'], col)
            cn = cn + 1

        # 出院诊断
        if '出院记录' in item:
            sheet.cell(rn, cn).value = get_json_value(item['出院记录'], '出院诊断')
        cn = cn + 1

        # 超声
        if '超声' in item:
            for item_cs in item['超声']:
                for item_cs_sj in item_cs['数据']:
                    arr = item_cs_sj.split(',')
                    if len(arr) == 3:
                        arr.append('')
                    # 检查部位
                    sub_cn = check_cs_part(arr[1])
                    str_cs = '影像描述：%s\n影像结论：%s\n\n' % (arr[2], arr[3])
                    if check_cell_empty(sheet.cell(rn, cn + sub_cn).value):
                        sheet.cell(rn, cn + sub_cn).value = str_cs
        cn = cn + len(columns[2])

        # 放射
        fs_str_dict = {'CT': [], 'MR': [], 'DR': [], 'X线': []}
        cn_start_ct, cn_start_mr = cn, cn + len(columns[3])
        cn_start_dr, cn_start_xx = cn_start_mr + len(columns[4]), cn_start_mr + len(columns[4]) + len(columns[5])
        if '放射' in item:
            for item_fs in item['放射']:
                for item_fs_sj in item_fs['数据']:
                    arr = item_fs_sj.split(',')
                    if len(arr) == 3:
                        arr.append('')
                    type = arr[0].upper()
                    part = arr[1]
                    str_fs = '影像表现：%s\n影像诊断：%s\n\n' % (arr[2], arr[3])
                    if 'CT' in type:
                        sub_cn_arr = check_ct_parts(part)
                        for sub_cn in sub_cn_arr:
                            if check_cell_empty(sheet.cell(rn, cn_start_ct + sub_cn).value):
                                sheet.cell(rn, cn_start_ct + sub_cn).value = str_fs
                    elif 'MR' in type:
                        sub_cn_arr = check_mr_parts(part)
                        for sub_cn in sub_cn_arr:
                            if check_cell_empty(sheet.cell(rn, cn_start_mr + sub_cn).value):
                                sheet.cell(rn, cn_start_mr + sub_cn).value = str_fs
                    elif 'DR' in type or 'CR' in type:
                        sub_cn_arr = check_dr_parts(part)
                        for sub_cn in sub_cn_arr:
                            if check_cell_empty(sheet.cell(rn, cn_start_dr + sub_cn).value):
                                sheet.cell(rn, cn_start_dr + sub_cn).value = str_fs
                    elif 'X线' in type:
                        sub_cn_arr = check_xx_parts(part)
                        for sub_cn in sub_cn_arr:
                            if check_cell_empty(sheet.cell(rn, cn_start_xx + sub_cn).value):
                                sheet.cell(rn, cn_start_xx + sub_cn).value = str_fs

        cn = cn_start_xx + len(columns[6])

        # 病理
        str_bl = ''
        if '病理' in item:
            for item_bl in item['病理']:
                for item_bl_sj in item_bl['数据']:
                    arr = item_bl_sj.split(',')
                    str_bl = str_bl + '检查所见：%s\n诊断意见：%s\n\n' % (arr[0], arr[1])
        sheet.cell(rn, cn).value = str_bl
        cn = cn + 1

        # 检验
        str_dict, cols_jy = {}, columns[8]
        cols_jy_set = set(cols_jy)
        if '检验' in item:
            for item_jy in item['检验']:
                for item_jy_sj in item_jy['数据']:
                    arr = item_jy_sj.split(',')
                    key = arr[2]
                    if key in cols_jy_set:
                        if key not in str_dict:
                            str_dict[key] = ''
                        if str_dict[key] == '':
                            str_dict[key] = ','.join(arr[-3:])

        for col in cols_jy:
            if col in str_dict:
                sheet.cell(rn, cn).value = str_dict[col]
            cn = cn + 1

        # # 外院
        # if re.search('(医院)|(外院)', item['入院记录']['现病史']):
        #     sheet.cell(rn, cn).value = 1


    workbook.save(r'data/%s/split_%s.xlsx' % (data_type, postfix))
