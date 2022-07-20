from openpyxl import load_workbook
import json


# wb = load_workbook(r'data\人机大赛病历库——规范化中.xlsx')
# sheet = wb['Sheet2']
#
# col, header = 17, '检查部位'    # US
# col, header = 22, '检查名称'    # CT
# # col, header = 24, '检查名称'    # MR
# # col, header = 25, '检查名称'    # CR/DR、移动DR
# # col, header = 26, '检查部位'    # ERCP
# rn = 3
#
# results = []
#
# for i in range(3, sheet.max_row + 1):
#     if sheet.cell(i, 1).value == None:
#         break
#
#     val = sheet.cell(i, col).value
#     if val != None and val != '':
#         arr = val.split('\n')
#
#     for l in arr:
#         if header in l:
#             results.append(l.split('：')[1])
#
# results = sorted(list(set(results)))
#
# for r in results:
#     print(r)


if __name__ == '__main__':
    # postfix = '1432'
    postfix = '2724'
    # postfix = '1611'

    # 加载json数据
    json_data = ''
    with open(r'data/汇总结果_%s.json' % postfix) as f:
        json_data = json.load(f, strict=False)

    # results = []
    # for record in json_data:
    #     if '超声' in record:
    #         for item in record['超声']:
    #             for line in item['数据']:
    #                 arr = line.split(',')
    #                 results.append(arr[1])
    # results = sorted(list(set(results)))
    # for r in results:
    #     print(r)

    results_dict = {}
    for record in json_data:
        if '放射' in record:
            for item in record['放射']:
                for line in item['数据']:
                    arr = line.split(',')
                    if arr[0] not in results_dict:
                         results_dict[arr[0]] = []
                    results_dict[arr[0]].append(arr[1])


    for key in results_dict.keys():
        results_dict[key] = sorted(list(set(results_dict[key])))
    for key in results_dict.keys():
        print('')
        print('')
        print('-------------------------')
        print(key)
        for r in results_dict[key]:
            print(r)
