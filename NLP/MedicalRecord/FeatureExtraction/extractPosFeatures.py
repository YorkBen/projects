from openpyxl import load_workbook, Workbook

cols = []
results = []
with open('tmp.txt') as f:
    for idx, line in enumerate(f.readlines()):
        arr = line.strip().split('	')
        if idx == 0:
            cols = arr
        else:
            # row_result = []
            # for ridx, e in enumerate(arr):
            #     if e == '1':
            #         row_result.append(cols[ridx])
            #     elif e == '3':
            #         row_result.append(cols[ridx] + '？')
            # results.append(row_result)
            results.append([int(e) for e in arr])


# #### 写excel
# wb = Workbook()
# sheet = wb.create_sheet('Sheet1', 0)
#
# # 写表头
# for ridx, row in enumerate(results):
#     for cidx, e in enumerate(row):
#         sheet.cell(ridx + 1, cidx + 1).value = e
#
# wb.save('tmp.xlsx')
# wb.close()

# 写txt
with open('tmp_r.txt', 'w') as f:
    for r in results:
        f.write('%s\n' % str(r))
