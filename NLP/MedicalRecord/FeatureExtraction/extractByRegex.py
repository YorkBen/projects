from openpyxl import load_workbook
import json
import re
from Lib.Utils import Utils

utils = Utils()

# 内部连续通用正则
inner_neg = '[^，；、无未不]{,4}'
inner_neg_l = '[^，；、无未不]{,8}'
inner = '[^，；、]*?'

features = [
    # {'id':'FSTX1', 'name': '放射痛（右肩、肩胛和背部）', 'src': ['现病史'], 'regex':'放射痛?.*((右肩)|(肩胛)|(背部))'},
    # {'id':'SFTX1', 'name': '上腹痛', 'src': ['现病史'], 'regex':'上腹.*痛'},
    # {'id':'YSX1', 'name': '厌食', 'src': ['现病史'], 'regex':'', 'type': 'model'},   # 饮食正常
    # {'id':'FXX1', 'name': '腹泻', 'src': ['现病史'], 'regex':'', 'type': 'model'},   # 大小便正常
    # {'id':'PQTX1', 'name': '盆腔疼痛', 'src': ['现病史'], 'regex':'盆腔.*痛'},
    # {'id':'BJQX1', 'name': '迅速波及全腹', 'src': ['现病史'], 'regex':''},
    # {'id':'HBX1', 'name': '黑便', 'src': ['现病史'], 'regex':''},
    # {'id':'YXFTX1', 'name': '右下腹疼痛', 'src': ['现病史'], 'regex':'右下腹.*痛'},
    # {'id':'ZYXTX1', 'name': '转移性右下腹疼痛', 'src': ['现病史'], 'regex':'', 'type': 'model'},
    # {'id':'ZFTX1', 'name': '中腹痛', 'src': ['现病史'], 'regex':'中腹.*痛'},
    # {'id':'YSFTX1', 'name': '右上腹痛', 'src': ['现病史'], 'regex':'右上腹.*痛'},
    # {'id':'FSDFX1', 'name': '放射痛（放射到同侧半腹或背部）', 'src': ['现病史'], 'regex':''},
    # {'id':'CXFTX1', 'name': '持续性腹痛', 'src': ['现病史'], 'regex':'((腹|(脐周)|(剑突下)).*?痛.*?[^每次]持续)|(持续.*?(腹|(脐周)|(剑突下)).*?痛)|((腹|(脐周)|(剑突下)).*?持续.*?痛)', 'negregex': '((腹|(脐周)|(剑突下)).*?痛.*?((间断)|(阵发)|(间歇)))|(((间断)|(阵发)|(间歇)).*?(腹|(脐周)|(剑突下)).*?痛)|((腹|(脐周)|(剑突下)).*?((间断)|(阵发)|(间歇)).*?痛)'},
    # {'id':'YHJZX1', 'name': '仰卧加重，坐位缓解', 'src': ['现病史'], 'regex':''},
    # {'id':'ZFXTX1', 'name': '阵发性疼痛', 'src': ['现病史'], 'regex':'阵发性'},
    # {'id':'JLFTX1', 'name': '剧烈腹痛', 'src': ['现病史'], 'regex':'(腹.*剧.*痛)|(剧.*腹.*痛)|(腹.*痛.*剧)'},
    # {'id':'TZPBX1', 'name': '停止排便排气', 'src': ['现病史'], 'regex':'', 'type': 'model'},
    # {'id':'YDLXX1', 'name': '阴道流血', 'src': ['现病史'], 'regex':'阴道.*流血'},
    # {'id':'PNGBX1', 'name': '排尿改变', 'src': ['现病史'], 'regex':'', 'type': 'model'},
    # {'id':'FTSCFX1', 'name': '放射痛（侧腹、腹股沟、睾丸或大阴唇）', 'src': ['现病史'], 'regex':'', 'type': 'model'},
    # {'id':'YCFTX1', 'name': '腰/侧腹痛', 'src': ['现病史'], 'regex':'', 'type': 'model'},
    #

    {'id':'XKT1', 'name': '失血性休克', 'src': ['体格检查'], 'regex':'', 'default': 0, 'type': 'model'}, #
    {'id':'MYXJ1', 'name': '自身免疫性疾病', 'src': ['既往史'], 'regex':'', 'default': 0, 'type': 'model'}, # 模型
    {'id':'YXZLJ1', 'name': '胰腺肿瘤、囊肿病史', 'src': ['既往史'], 'regex':'胰腺.*((肿瘤)|(囊肿))', 'default': 0, 'type': 'model'}, # 模型
    {'id':'JWSJ1', 'name': '既往生殖器感染', 'src': ['既往史'], 'regex':'', 'default': 0, 'type': 'model'}, # 模型?
    {'id':'FBSS1', 'name': '腹部手术史', 'src': ['既往史', '手术外伤史'], 'regex':'腹' + inner_neg + '手?术', 'default': 0, 'type': 'model'},    # 模型
    {'id':'FBZJ1', 'name': '腹部肿瘤史', 'src': ['既往史', '手术外伤史'], 'regex':'腹部' + inner_neg + '肿?瘤', 'default': 0, 'type': 'model'},   # 模型
    {'id':'HLSJ1', 'name': '肿瘤化疗史', 'src': ['既往史', '手术外伤史'], 'regex':'化疗', 'default': 0, 'type': 'model'}, # 模型？

    #
    # {'id':'TJX1', 'name': '停经', 'src': ['现病史'], 'regex':'', 'type': 'model'},

    {'id':'TYX1', 'name': '头晕', 'src': ['现病史'], 'regex':'头晕',  'default': 0},
    {'id':'FZX1', 'name': '腹胀', 'src': ['现病史', '查体'], 'regex':'([腹饱]胀)|(胀[痛疼])|(腹' + inner_neg + '胀)',  'default': 0},
    {'id':'EXOTX1', 'name': '恶心呕吐', 'src': ['现病史'], 'regex':'(恶心)|(呕吐)|呕|吐',  'default': 0},
    {'id':'TRFZX1', 'name': '突然发作', 'src': ['现病史'], 'regex':'(突发)|(突然)', 'default': 0},
    {'id':'JTX1', 'name': '绞痛', 'src': ['现病史'], 'regex':'绞痛', 'default': 0},
    {'id':'HDX1', 'name': '黄疸', 'src': ['现病史'], 'regex':'(黄疸)|(黄染)|(发黄)', 'default': 0},
    {'id':'ZLBX1', 'name': '坐立不安', 'src': ['现病史'], 'regex':'坐立不安', 'default': 0},


    {'id':'YWJ1', 'name': '异物摄入史', 'src': ['既往史', '现病史'], 'regex':''}, # 没有查询词
    {'id':'SLGJ1', 'name': '输卵管积水', 'src': ['现病史', '既往史'], 'regex':''}, # 0
    {'id':'YNJ1', 'name': '饱餐、进食油腻食物史', 'src': ['现病史', '既往史'], 'regex':'(饱餐)|(油腻)|(油脂)', 'default': 0}, # 1
    {'id':'JSJ1', 'name': '禁食', 'src': ['现病史', '既往史'], 'regex':'', 'default': 0}, # 0
    {'id':'CXCRJ1', 'name': 'CXCR1低表达', 'src': ['现病史', '既往史'], 'regex':'CXCR1' + inner_neg + '低'}, # 0
    {'id':'ZZJJ1', 'name': '重症监护室（应激性溃疡）', 'src': ['现病史', '既往史'], 'regex':'应激性溃疡', 'default': 0}, # 0
    {'id':'GYJ1', 'name': '高盐饮食', 'src': ['现病史', '既往史'], 'regex':'高盐'}, # 0
    {'id':'GCJ1', 'name': '高草酸饮食', 'src': ['现病史', '既往史'], 'regex':'高草酸'}, # 0
    {'id':'YSSJ1', 'name': '饮水较少', 'src': ['现病史', '既往史'], 'regex':'(饮水少)|(饮水较少)|(少饮水)'}, # 0
    {'id':'XHDJ1', 'name': '消化道溃疡', 'src': ['现病史', '既往史'], 'regex':'(肠溃疡)|(十二指肠球部溃疡)|(胃溃疡)|(消化道溃疡)|(胃窦巨大溃疡)|(胃角巨大溃疡)', 'default': 0}, # 6
    {'id':'YGXJ1', 'name': '超过一个性伴侣', 'src': ['现病史', '既往史'], 'regex':'(冶游)|(不洁性交)', 'default': 0},  # 0
    {'id':'XBLJ1', 'name': '新的性伴侣', 'src': ['现病史', '既往史'], 'regex':'(冶游)|(不洁性交)', 'default': 0}, # 0
    {'id':'DGJ1', 'name': '低钙饮食', 'src': ['现病史', '既往史'], 'regex':'低钙饮食', 'default': 0}, # 0

    {'id':'HPJ1', 'name': 'Hp感染史', 'src': ['既往史'], 'regex':'(幽门螺旋杆菌感染)|(HP感染)|(HP（+）)|(HP阳性)', 'default': 0}, # 0 现病史？
    {'id':'PQYJ1', 'name': '盆腔炎病史', 'src': ['既往史'], 'regex':'', 'default': 0}, # 0
    {'id':'NLJ1', 'name': '尿路感染史', 'src': ['既往史'], 'regex':'(肾盂肾炎)|(肾炎)|(输尿管炎)|(膀胱炎)|(尿道炎)', 'default': 0}, # 0 现病史？
    {'id':'NSZJ1', 'name': '尿石症', 'src': ['既往史'], 'regex':'(肾结石)|(肾小结石)|(肾多发小结石)|(肾点状结石)|(输尿管结石)|(输尿管上段结石)', 'default': 0}, # 29
    {'id':'JTLJ1', 'name': '畸胎瘤', 'src': ['既往史'], 'regex':'', 'default': 0}, # 1 现病史？
    {'id':'WBSLJ1', 'name': '胃泌素瘤', 'src': ['既往史'], 'regex':'(胃泌素瘤)|(卓-艾综合征)|(Z-E综合征)', 'default': 0}, # 0 现病史？
    {'id':'LCPJ1', 'name': '卵巢旁囊肿', 'src': ['既往史'], 'regex':'', 'default': 0}, # 0
    {'id':'DJSJ1', 'name': '胆结石', 'src': ['既往史', '现病史'], 'regex':'(胆结石)|(胆管结石)|(胆总管结石)|(胆囊结石)|(胆石症)', 'default': 0}, # 现病史 ? 没有现病史130，有了34
    {'id':'MSLY1', 'name': '美沙拉明、速尿、氯沙坦、6-巯基嘌呤或硫唑嘌呤、异烟肼、袢利尿剂和去羟肌苷使用史',
        'src': ['既往史'], 'regex':'(美沙拉明)|(速尿)|(呋塞米)|(氯沙坦)|(科素亚)|(洛沙坦)|(罗沙藤)|(6-巯基嘌呤)|(乐疾宁)|(巯嘌呤)|(巯基嘌呤)|(硫唑嘌呤)|(依木兰)|(异烟肼)|(雷米封)|(袢利尿剂)|(去羟肌苷)'
        , 'default': 0}, # 3
    {'id':'GGXZJ1', 'name': '高钙血症', 'src': ['既往史'], 'regex':'', 'default': 0}, # 0
    {'id':'GYSZJ1', 'name': '高甘油三酯血症', 'src': ['既往史'], 'regex':'(高脂血症?)|(高血脂症?)|(((甘油三酯)|(血脂)|(胆固醇))' + inner_neg + '高)', 'default': 0}, # 3
    {'id':'BDJ1', 'name': '病毒(腮腺炎、柯萨奇病毒、巨细胞病毒、水痘、单纯疱疹病毒、HIV)、细菌(支原体、军团菌、钩端螺旋体、沙门氏菌)、寄生虫(弓形虫、隐孢子虫、蛔虫)和真菌(曲霉)',
        'src': ['既往史'], 'regex':'(腮腺炎)|(柯萨奇病毒)|(巨细胞病毒)|(水痘)|(疱疹)|(HIV)|(人[类体]?免疫缺陷病毒)|(艾滋病)|(支原体)|(军团菌)|(钩端螺旋体)|(沙门氏菌)|(弓形虫)|(隐孢子虫)|(蛔虫)|(真菌)|(曲霉)'
        , 'default': 0}, # 1
    {'id':'JYTBJ1', 'name': 'PRSS1、SPINK1、CFTR、CASR、CTRC基因突变', 'src': ['既往史'], 'regex':'(PRSS1)|(SPINK1)|(CFTR)|(CASR)|(CTRC)', 'default': 0}, # 0

    {'id':'YCBJ1', 'name': '炎症性肠病病史', 'src': ['既往史'], 'regex':'(炎症性肠病)|(溃疡性结肠炎)|(溃结)|(克罗恩)', 'default': 0},  # 0
    {'id':'QSYJ1', 'name': '憩室炎', 'src': ['既往史'], 'regex':'憩室炎', 'default': 0}, # 5
    {'id':'CXMJ1', 'name': '肠系膜缺血', 'src': ['既往史'], 'regex':'肠系膜' + inner_neg + '缺血', 'default': 0}, # 0
    {'id':'GCGJ1', 'name': '钩虫感染', 'src': ['既往史'], 'regex':'钩虫', 'default': 0}, # 0
    {'id':'FZTJ1', 'name': '非甾体抗炎药，氯化钾肠溶片使用史', 'src': ['既往史'],
        'regex':'(非甾体抗炎药)|(阿司匹林)|(布洛芬)|(对乙酰氨基酚)|(吲哚美辛)|(萘普生)|(萘普酮)|(氯芬酸)|(尼美舒利)|(罗非昔布)|(塞来昔布)|(氯化钾)|(补达秀)',
        'default': 0}, # 7
    {'id':'WYJJ1', 'name': '二乙基己烯雌酚', 'src': ['既往史'], 'regex':'二乙基己烯雌酚', 'default': 0}, # 0
    {'id':'GNJJ1', 'name': '宫内节育器使用(超过2年)', 'src': ['既往史'], 'regex':'(节育器)|(节育环)', 'default': 0},    # 0
    {'id':'LGCJ1', 'name': '类固醇', 'src': ['既往史'], 'regex':'(大力补)|(康力龙)|(康复龙)|(睾酮)', 'default': 0},  # 0
    {'id':'HBSJ1', 'name': '环丙沙星、三硅酸镁、磺胺药物、氨苯蝶啶、茚地那韦、愈创甘油醚、麻黄碱、袢利尿剂（呋塞米）、碳酸酐酶抑制剂、泻药（开塞露）、阿昔洛韦、环利尿剂、乙酰唑胺、茶碱、糖皮质激素（泼尼松）、噻嗪、水杨酸、丙磺舒、别嘌呤醇服用史',
        'src': ['既往史'], 'regex':'(环丙沙星)|(三硅酸镁)|(磺胺)|(苯磺胺)|(氨苯蝶啶)|(茚地那韦)|(愈创甘油醚)|(麻黄碱)|(托拉塞米)|(布美他尼)|(呋塞米)|(乙酰唑胺)|(碳酸酐酶抑制剂)|(泻药)|(开塞露)|(阿昔洛韦)|(环利尿剂)|(乙酰唑胺)|(茶碱)|(泼尼松)|(甲泼尼松龙)|(倍他米松)|(氢化可的松)|(可的松)|(地塞米松)|(噻嗪)|(氢氯噻嗪)|(阿司匹林)|(水杨酸)|(丙磺舒)|(别嘌呤醇)',
        'default': 0}, # 26
    {'id':'SJJJ1', 'name': '杀精剂接触史', 'src': ['既往史'], 'regex':'壬苯醇醚', 'default': 0}, # 1

    {'id':'TNBJ1', 'name': '糖尿病', 'src': ['既往史'], 'regex':'', 'default': 0}, # 20
    {'id':'GXYJ1', 'name': '高血压', 'src': ['既往史'], 'regex':'', 'default': 0}, # 29
    {'id':'TFJ1', 'name': '痛风', 'src': ['既往史'], 'regex':'', 'default': 0}, # 3
    {'id':'FPJ1', 'name': '肥胖', 'src': ['既往史'], 'regex':'', 'default': 0}, # 0
    {'id':'JJJ1', 'name': '结节病', 'src': ['既往史'], 'regex':'', 'default': 0}, # 0
    {'id':'KLEJ1', 'name': '克罗恩病', 'src': ['既往史'], 'regex':'克罗恩病?', 'default': 0}, # 1
    {'id':'JZPJ1', 'name': '原发性甲状旁腺功能亢进', 'src': ['既往史'], 'regex':'甲状旁腺功能亢进', 'default': 0}, # 2
    {'id':'JKJ1', 'name': '甲亢', 'src': ['既往史'], 'regex':'(甲状腺功能亢进)|(甲亢)', 'default': 0}, # 10
    {'id':'DFXJ1', 'name': '多发性骨髓瘤', 'src': ['既往史'], 'regex':'多发性骨髓瘤', 'default': 0}, # 0
    {'id':'SXGJ1', 'name': '肾小管酸中毒', 'src': ['既往史'], 'regex':'肾小管酸中毒', 'default': 0}, # 0
    {'id':'MXSJ1', 'name': '慢性肾病', 'src': ['既往史'], 'regex':'(尿毒症)|(肾衰)', 'default': 0}, # 13

    {'id':'DGZYJ1', 'name': '内镜逆行胰胆管造影、EUS与FNA、主动脉手术胰腺切除术史', 'src': ['既往史', '手术外伤史'], \
                'regex':'(内镜逆行胰胆管造影)|(ERCP)|(EUS-FNA)|(超声内镜引导下细针穿刺)|(EUS)|(FNA)|(主动脉' + inner_neg + '术)|(胰腺' + inner_neg + '((切除)|术))', 'default': 0},


    {'id':'NMYWJ1', 'name': '子宫内膜异位症', 'src': ['既往史', '手术外伤史'], 'regex':'((子宫内膜异位)|(内异症))', 'default': 0},
    {'id':'SQJ1', 'name': '疝气或疝气修复史', 'src': ['既往史', '手术外伤史'], 'regex':'[^食管裂孔]{4}疝', 'default': 0}, # 3
    {'id':'LWYJ1', 'name': '阑尾炎', 'src': ['既往史', '手术外伤史'], 'regex':'((阑尾炎)|(阑尾脓肿))', 'default': 0}, # 不带切除3
    {'id':'SLRJ1', 'name': '输卵管手术史', 'src': ['既往史', '手术外伤史'], 'regex':'输卵管' + inner_neg + '((手?术)|(结扎))', 'default': 0},
    {'id':'WRDJ1', 'name': '胃绕道手术史、减肥手术史', 'src': ['既往史', '手术外伤史'], 'regex':'(胃绕道)|(减肥手术)|(缩胃)|(胃缩)', 'default': 0}, #全是2
    {'id':'DCZJ1', 'name': '短肠综合征', 'src': ['既往史', '手术外伤史'], 'regex':'短肠综合征', 'default': 0}, # 0

    {'id':'RSSJ1', 'name': '既往异位妊娠史', 'src': ['既往史', '手术外伤史'], 'regex':'异位妊娠', 'default': 0},

    {'id':'BZFS1', 'name': '板状腹', 'src': ['查体', '专科情况（体检）'], 'regex':'(板状腹)|(腹肌紧张)|(腹肌强直)|(腹壁紧张)|(腹壁强直)'},

    {'id':'LWYJJ1', 'name': '阑尾炎家族史阳性', 'src': ['家族史'], 'regex':'(阑尾炎)|(阑尾脓肿)', 'default': 0}, # 全是2
    {'id':'NJJ1', 'name': '尿石症家族史', 'src': ['家族史'], 'regex':'(肾结石)|(肾小结石)|(肾多发小结石)|(肾点状结石)|(输尿管结石)|(输尿管上段结石)', 'default': 0}, # 5

    {'id':'FYTS1', 'name': '下腹压痛', 'src': ['查体'], 'regex':'[下全]腹' + inner_neg + '压痛', 'negregex': '([下全]腹' + inner + '(无|(未见)|(未及))' + inner + '压痛)|(上腹' + inner + '压痛)'},
    {'id':'ZGYTS1', 'name': '子宫压痛', 'src': ['查体'], 'regex':'子宫' + inner_neg + '压痛'}, #全是2
    {'id':'FJTS1', 'name': '附件区压痛', 'src': ['查体'], 'regex':'(附件区' + inner_neg + '压痛)|(输卵管压痛)|(卵巢压痛)'}, #全是2
    {'id':'FJYTS1', 'name': '附件压痛', 'src': ['查体'], 'regex':'(附件区' + inner_neg + '压痛)|(输卵管压痛)|(卵巢压痛)'}, # 全是2
    {'id':'SZKTS1', 'name': '肾脏叩击痛', 'src': ['查体'], 'regex':'肾' + inner_neg + '叩击痛', 'negregex': '肾' + inner + '(无|(未见)|(未及))' + inner + '叩击痛'},
    {'id':'CDNZS1', 'name': '可能触到囊肿', 'src': ['查体'], 'regex':'触?[及到]' + inner_neg_l + '囊肿'}, # 全是2
    {'id':'GJCJS1', 'name': '宫颈刺激', 'src': ['查体'], 'regex':'宫颈' + inner_neg + '((刺激)|(举痛)|(摇举痛))'},  # 全是2

    {'id':'KZGYT1', 'name': '腹部叩诊鼓音', 'src': ['体格检查', '查体'], 'regex':'(腹部)?(叩诊)?' + inner_neg + '鼓音'},
    {'id':'CYXSS1', 'name': '肠鸣音消失', 'src': ['体格检查', '查体'], 'regex':'肠鸣音消失'},
    {'id':'CYKT1', 'name': '肠鸣音亢进（病程早期）', 'src': ['体格检查', '查体'], 'regex':'肠鸣音亢进'},
    {'id':'CYRT1', 'name': '肠鸣音减弱（病程晚期）', 'src': ['体格检查', '查体'], 'regex':'肠鸣音.?[弱低]'},

    {'id':'HMT1', 'name': '昏迷', 'src': ['体格检查', '专科情况（体检）', '查体'],
       'regex':'(昏迷)|(意识不清)|(呼之不应)|(意识丧失)|(随意运动消失)|(对外界的刺激的反应迟钝或丧失)|(瞳孔散大)|(对光反射消失)|(双侧瞳孔不等大)|(神经反射消失)', 'default': 0},
    {'id':'TWGT1', 'name': '体温升高', 'src': ['体格检查'], 'regex':''},
    {'id':'MBKT1', 'name': '脉搏显著加快', 'src': ['体格检查'], 'regex':''},
    {'id':'DXYT1', 'name': '低血压', 'src': ['体格检查'], 'regex':''},
    {'id':'JJZT1', 'name': '反跳痛或肌紧张', 'src': ['体格检查', '现病史', '查体'], 'regex':'(反跳痛)|(肌紧张)|(腹' + inner_neg + '紧)|(腹肌抵触感)|(板状腹)|(腹强直)'},
    {'id':'MFST1', 'name': '墨菲征Murphy征', 'src': ['体格检查', '查体'], 'regex':'((墨菲征)|(Murphys?征?：?’?))', 'negregex': '((墨菲征)|(Murphys?))' + inner + '阴性'},
    {'id':'ZDDNT1', 'name': '肿大胆囊', 'src': ['体格检查'], 'regex':'上腹' + inner_neg + '肿块', 'default': 0}, #
    {'id':'FJT1', 'name': '附件肿块', 'src': ['体格检查'], 'regex':'(上腹' + inner_neg + '肿块)|(附件' + inner_neg + '触及' + inner_neg + '[肿包]块)|(附件' + inner_neg + '增[粗厚])', 'default': 0}, #


    {'id':'XJG1', 'name': '酗酒', 'src': ['个人史'], 'regex':'酗酒'},
    {'id':'XYGS1', 'name': '吸烟', 'src': ['个人史'], 'regex':'吸烟'},

    {'id':'NL35', 'name': '年龄35岁及以上', 'src': ['年龄'], 'regex':''},

    {'id':'LCJ1', 'name': '既往流产（包括人工流产）', 'src': ['婚育史'], 'regex':'流产'},
    {'id':'BYHJ1', 'name': '不孕(风险随不孕时间的延长而增加)', 'src': ['婚育史'], 'regex':'不孕'}
]


def load_medical_data(path):
    """
    加载病历结构化数据
    """
    # path = r'data/汇总结果.json'
    json_data = ''
    with open(path) as f:
        json_data = json.load(f, strict=False)

    return json_data

def load_feature_sheet(workbook_path, sheet_name):
    """
    workbook_path: excel路径
    sheet_name：表单名
    """
    # workbook_path = r"data/高发病率腹痛疾病特征标注2022.6.23.xlsx"
    # sheet_name = "前500个疾病特征标注"
    workbook = load_workbook(workbook_path)
    sheet = workbook[sheet_name]

    cols, rows, sheet_data = [''], [''], {}
    # 写入列名
    for j in range(1, sheet.max_column + 1):
        if sheet.cell(1, j).value is None:
            break

        cols.append(sheet.cell(1, j).value)

    # 写入行名，及数据
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(i, 1).value is None:
            break

        key = str(sheet.cell(i, 1).value)
        rows.append(key)
        sheet_data[key] = {}
        for j in range(2, len(cols)):
            # 第2列是金标准疾病
            if j == 2:
                sheet_data[key][cols[j]] = sheet.cell(i, j).value
            elif sheet.cell(i, j).value is not None:
                sheet_data[key][cols[j]] = int(sheet.cell(i, j).value)
            else:
                sheet_data[key][cols[j]] = 0

    return sheet_data

def split_text(text):
    # text = text.replace('，间断', ' 间断').replace('，持续', ' 持续').replace('，阵发', ' 阵发')
    # text = text.replace('，呈', ' 呈').replace('，为', ' 为')
    # text = text.replace('，', '。')
    return text.split('。')

def search_by_regex(text, regex, negregex, default):
    pos_match, pneg_match, neg_match, rt1, rt2 = None, None, None, 2, 2
    for t in split_text(text):
        match1 = re.search(regex, t)
        pos_match = match1
        if match1:
            regex2 = '[^,，；;！]*(' + regex + ')[^,，；;！]*'
            match11 = re.search(regex2, t)
            t_ = t[match11.span()[0]:match11.span()[1]]

            # 否定
            match1 = re.search(regex, t_)
            mt1_sp2 = match1.span()[1]
            match2 = re.search(r"(无[^痛])|(不[^详全])|未|(否认)", t_)
            match3 = re.search(r"(不明显)|(阴性)|(排除)|([(（][-—][)）])", t_)
            if match2 and match2.span()[1] < mt1_sp2 and not '诱因' in t_[:mt1_sp2]:
                pneg_match, rt1 = match2, 0
            elif match3 and match3.span()[0] >= mt1_sp2:
                pneg_match, rt1 = match3, 0
            else:
                rt1 = 1

            # # 待排、可能
            # match4 = re.search("(待排)|(可能)|[?]|？", t_)
            # if match4 and match4.span()[0] >= mt1_sp2 and (match4.span()[0] - mt1_sp2) <= 2:
            #     pneg_match, rt1 = match4, 3

            break

    if negregex is not None:
        for t in split_text(text):
            match1 = re.search(negregex, t)
            if match1:
                neg_match, rt2 = match1, 0
                break

    if pos_match is not None and neg_match is not None:
        if pos_match.span()[1] - pos_match.span()[0] <= neg_match.span()[1] - neg_match.span()[0]:
            return rt1, pos_match, pneg_match
        else:
            return rt2, neg_match, ''
    elif pos_match is not None:
        return rt1, pos_match, pneg_match
    elif neg_match is not None:
        return rt2, neg_match, ''
    elif default is not None:
        return default, '', ''
    else:
        return 2, '', ''

def search_by_strc(record, src, name):
    """
    通过结构化数据查找
    """
    if name == '肿大胆囊':
        text = record['入院记录']['体格检查']['腹部']['胆囊']
        if text.startswith('未触及'):
            return 0
        else:
            return 1
    elif name == '腹部叩诊鼓音':
        text = record['入院记录']['体格检查']['腹部']['叩诊']
        if '鼓音' in text:
            return 1
        else:
            return 0
    elif name == '肠鸣音亢进（病程早期）':
        text = record['入院记录']['体格检查']['腹部']['肠鸣音']
        if '亢进' in text:
            return 1
        else:
            return 0
    elif name == '肠鸣音减弱（病程晚期）':
        text = record['入院记录']['体格检查']['腹部']['肠鸣音']
        if '弱' in text:
            return 1
        else:
            return 0
    elif name == '肠鸣音消失':
        text = record['入院记录']['体格检查']['腹部']['肠鸣音']
        if '消失' in text:
            return 1
        else:
            return 0
    elif name == '失血性休克':    #医学生进一步讨论完成
        text = record['入院记录']['体格检查']['腹部']['肠鸣音']
        return 2
    elif name == '昏迷':
        text = record['入院记录']['体格检查']['一般情况']['神志']
        if '昏迷' in text:
            return 1
        else:
            return 0
    elif name == '体温升高':
        text = utils.format_temprature(record['入院记录']['体格检查']['生命体征']['体温']).replace('℃', '')
        if text != '':
            if float(text) > 37.5:
                return 1
            else:
                return 0
        else:
            return 2
    elif name == '脉搏显著加快':
        text = utils.format_num(record['入院记录']['体格检查']['生命体征']['脉搏'])
        if text != '':
            num = int(text)
            if num > 100 and num < 1000:    # 数据错误8979，应该是89，如果>10000，应该是> 100
                return 1
            elif num > 10000:
                return 1
            else:
                return 0
        else:
            return 2
    elif name == '低血压':
        text = utils.format_pressure(record['入院记录']['体格检查']['生命体征']['血压'])
        arr = text.split('/')
        if len(arr) != 2:
            return 2

        ssy = utils.format_num(arr[0])
        szy = utils.format_num(arr[1])
        if ssy == '' or szy == '':
            return 2
        else:
            if int(ssy) < 90 or int(szy) < 60:
                return 1
            else:
                return 0
    elif name == '反跳痛或肌紧张':
        text = record['入院记录']['体格检查']['腹部']['反跳痛']
        if text == '':
            return 2
        elif '无' in text:
            return 0
        else:
            return 1
    elif name == '墨菲征Murphy征':
        text = record['入院记录']['体格检查']['腹部']['Murphy征'].strip()
        if text == '':
            return 2
        elif '无' in text or text == '-':
            return 0
        else:
            return 1
    elif name == '附件肿块':    #医学生进一步讨论完成
        text = record['入院记录']['体格检查']['腹部']['腹部包块'].strip()
        if text.startswith('未触及'):
            return 0
        else:
            if '附件' in text:
                return 1
            else:
                return 0
    elif name == '吸烟':
        text = record['入院记录']['个人史']['吸烟史']
        if text.replace('-', '') == '':
            return 2
        elif record['入院记录']['个人史']['戒烟时间'] != '':
            return 1
        elif '无' in text:
            return 0
        else:
            return 1
    elif name == '酗酒':
        text = record['入院记录']['个人史']['饮酒史']
        if text.replace('-', '') == '':
            return 2
        elif '无' in text:
            return 0
        else:
            return 1
    elif name == '年龄35岁及以上':
        text = record['入院记录']['年龄']
        age = utils.format_by_type(text, 'age')
        if age != '':
            if int(age) >= 35:
                return 1
            else:
                return 0
        else:
            return 2
    elif name == '既往流产（包括人工流产）':
        num1 = utils.format_num(record['入院记录']['婚育史']['自然流产'])
        num2 = utils.format_num(record['入院记录']['婚育史']['人工流产'])
        if num1 == '' and num2 == '' and record['入院记录']['婚育史']['产'] == '':
            return 2
        else:
            num1 = int(num1) if num1 != '' else 0
            num2 = int(num2) if num2 != '' else 0
            if num1 + num2 > 0:
                return 1
            else:
                return 0
    elif name == '不孕(风险随不孕时间的延长而增加)':
        if '未婚' in record['入院记录']['婚育史']['婚育史']:
            return 2
        else:
            has_num = False
            if utils.format_num(record['入院记录']['婚育史']['妊娠'], default='0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['产'], default='0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['自然生产'], default='0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['手术生产'], '0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['自然流产'], '0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['人工流产'], '0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['早产'], '0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['死产'], '0') != '0':
                has_num = True
            if utils.format_num(record['入院记录']['婚育史']['引产'], '0') != '0':
                has_num = True

            if has_num:
                return 0
            else:
                mar_age = utils.format_age(record['入院记录']['婚育史']['结婚年龄'])
                age = utils.format_age(record['入院记录']['年龄'])
                if mar_age != '' and age != '':
                    # 数据空的认为缺失
                    if int(age) - int(mar_age) > 1 and utils.format_num(record['入院记录']['婚育史']['产']) != '':
                        return 1
                return 2


def get_text_from_mr(record, key):
    """
    从一条病历结构化文本中拼接文本
    """
    text = ''
    if key in record:
        text = text + str(record[key]) + '。'
    elif key == '既往史':
        text = text + str(record['入院记录']['病史小结']['既往史']).replace('，', '、') + '。'
        text = text + str(record['入院记录']['既往史']) + '。'
    elif '首次病程' in record and key == '查体':
        text = text + str(record['首次病程']['病例特点']['查体']) + '。'
        text = text + str(record['首次病程']['诊断依据']) + '。'
    elif key in record['入院记录']:
        text = text + str(record['入院记录'][key]) + '。'

    return text

def process_by_regular(json_data, mrnos=None):
    """
    病历数据抽取正则部分特征
    json_data：病历数据
    mrnos：医保号码
    """
    results = {}
    for record in json_data:
        # 如果指定要处理哪些医保编号，这里进行一下过滤
        if mrnos is not None and record["医保编号"] not in mrnos:
            continue

        results[record["医保编号"]] = {}
        for feature in features:
            # 用模型处理的特征跳过
            if 'type' in feature and feature['type'] == 'model':
                continue

            # 抓取来源文本
            text, result_strc = '', -1  # 来源病历文本，按照结构化查找结果
            for src in feature['src']:
                if src in ['体格检查', '个人史', '年龄', '婚育史']:
                    result_strc = search_by_strc(record, src, feature['name'])
                else:
                    text = text + get_text_from_mr(record, src)

            # 按结构查找
            if text == '' and result_strc == -1:
                result, match1, match2 = 2, '', ''
            elif text == '' or result_strc == 1:
                result, match1, match2 = result_strc, '', ''
            else:
                regex = feature['regex'] if feature['regex'] != '' else feature['name']
                negregex = feature['negregex'] if 'negregex' in feature else None
                result_, match1, match2 = search_by_regex(text, regex, negregex, feature['default'] if 'default' in feature else 0)
                if result_ == 0 or result_ == 1 or result_strc == -1:
                    result = result_

            results[record["医保编号"]][feature['id']] = (result, match1, match2, record)

    return results

def stat_results(results, labeled_data, ignore02=True):
    """
    将正则提取结果和人工标记结果做比对，并统计输出
    输入：
        1. results： "医保编号" -> 特征id -> (result_class, match1, match2, record)
        2. labeled_data："医保编号" -> 特征id -> labeled_class
        3. ignore02：当人工标记和正则提取标志分别未0和2时，是否区分结果是否相同。
    """
    stat_features = {}
    for mrno in results.keys():
        for feature in features:
            # 用模型处理的特征跳过
            if 'type' in feature and feature['type'] == 'model':
                continue

            fid, fname = feature['id'], feature['name']
            if fid in labeled_data[mrno]:
                if fname not in stat_features:
                    stat_features[fname] = [0, 0]
                stat_features[fname][1] = stat_features[fname][1] + 1
                rno, lno = results[mrno][fid][0], labeled_data[mrno][fid]
                if rno == lno:
                    stat_features[fname][0] = stat_features[fname][0] + 1
                elif ignore02 and ((rno == 0 and lno == 2) \
                    or (rno == 2 and lno == 0)):
                    stat_features[fname][0] = stat_features[fname][0] + 1

    stat_results = [[feature, stat_features[feature][0], stat_features[feature][1]] for feature in stat_features.keys()]
    stat_results = sorted(stat_results, key=lambda x: x[1], reverse=True)
    for line in stat_results:
        print(line)


def merge_results(results, labeled_data, out_file, ignore02=True):
    """
    将正则提取结果和人工标记结果做比对，有差异部分用正则覆盖
    输入：
        1. results： "医保编号" -> 特征id -> (result_class, match1, match2, record)
        2. labeled_data："医保编号" -> 特征id -> labeled_class
        3. ignore02：当人工标记和正则提取标志分别未0和2时，是否区分结果是否相同。
    """
    for mrno in results.keys():
        for feature in features:
            # 用模型处理的特征跳过
            if 'type' in feature and feature['type'] == 'model':
                continue

            fid, fname = feature['id'], feature['name']
            if fid in labeled_data[mrno]:
                rno, lno = results[mrno][fid][0], labeled_data[mrno][fid]
                if rno == lno:
                    pass
                else:
                    labeled_data[mrno][fid] = rno

    keys = []
    for mrno in labeled_data.keys():
        keys = list(labeled_data[mrno].keys())
        break

    labels, train_x = [], [[] for i in range(len(keys[1:]))]
    for mrno in labeled_data.keys():
        for idx, key in enumerate(keys):
            if key == '金标准':
                labels.append(labeled_data[mrno][key])
            else:
                train_x[idx-1].append(labeled_data[mrno][key])

    label_set = list(set(labels))
    labels_id = []
    for label in labels:
        labels_id.append(label_set.index(label))

    with open(out_file, 'w') as f:
        for k in range(len(labels)):
            line = '%s,%s' % (labels[k], labels_id[k])
            for j in range(len(train_x)):
                 line = line + ',' + str(train_x[j][k])
            f.write(line + '\n')

    return labels, labels_id, train_x


def debug_feature(results, labeled_data, fid, ignore02=True, mrno_spec=None):
    """
    打印单个特征调试信息
    输入：
        1. results： "医保编号" -> 特征id -> (result_class, match1, match2, record)
        2. labeled_data："医保编号" -> 特征id -> labeled_class
        3. fid：特征id
        4. ignore02：当人工标记和正则提取标志分别未0和2时，是否区分结果是否相同。
        5. mrno_spec：指定医保编号
    """
    not_equal_ct = 0
    stat_features = {}
    for mrno in results.keys():
        if mrno_spec is not None and mrno != mrno_spec:
            continue

        for feature in features:
            if feature['name'] == fid or feature['id'] == fid:
                fid, fname, src_list = feature['id'], feature['name'], feature['src']

        if fid in labeled_data[mrno]:
            rno, lno = results[mrno][fid][0], labeled_data[mrno][fid]
            if rno == lno:
                pass
            elif ignore02 and ((rno == 0 and lno == 2) \
                or (rno == 2 and lno == 0)):
                pass
            else:
                not_equal_ct = not_equal_ct + 1
                print(rno, lno)
                print(results[mrno][fid][1], results[mrno][fid][2])
                # 打印病历文本
                text = ''
                for src in src_list:
                    text = text + get_text_from_mr(results[mrno][fid][3], src)
                print(text)
                print('')

    print(not_equal_ct)


if __name__ == '__main__':
    medical_data = load_medical_data(r'data/汇总结果_1432.json')
    sheet_data = load_feature_sheet(r"data/高发病率腹痛疾病特征标注2022.6.23.xlsx", "前500个疾病特征标注")
    results = process_by_regular(medical_data, set(list(sheet_data.keys())))
    # stat_results(results, sheet_data)
    # debug_feature(results, sheet_data, '腹胀')
    data = merge_results(results, sheet_data, r'data\train_data_202207.txt')
    print(data)


#
