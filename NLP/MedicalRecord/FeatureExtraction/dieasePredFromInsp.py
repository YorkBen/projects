# pip install openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import re
import os
import argparse

from Lib.MRRecordUtil import *
from Lib.InspRule import InspRule

def get_json_value(item, key):
    if key not in item:
        return ''
    else:
        return str(item[key])

def load_regex():
    lines = []
    for filename in ['放射', '超声']:
        with open(r'data/regex/%s.txt' % filename) as f:
            for l in f.readlines():
                lines.append(l.strip())

    return '((' + ')|('.join(lines) + '))'


def get_mzwy_texts(item):
    """
    获取门诊外院中包含实验室数据的文本，以。分割，返回句子数组
    """
    s1 = get_json_value(item['入院记录'], '门诊及院外重要辅助检查') if '入院记录' in item else ''
    s2 = get_json_value(item['入院记录']['病史小结'], '辅助检查') if '入院记录' in item and '病史小结' in item['入院记录'] else ''
    s3 = get_json_value(item['首次病程']['病例特点'], '辅助检查') if '首次病程' in item else ''
    s4 = get_json_value(item['出院记录']['入院情况'], '辅助检查') if '出院记录' in item else ''
    s5 = get_json_value(item['入院记录'], '现病史') if '入院记录' in item else ''

    isp = InspRule()
    regex = load_regex()
    texts = []
    for s in [s1, s2, s3, s4, s5]:
        for t in isp.split_text(s):
            if re.search(regex, t):
                texts.append(t)

    return texts

# def compare_sheet_data(sheet, diease, json_data):
#     cn_starts = [2, 7]
#     rn = 2
#     same_ct, total_ct = 0, 0
#     for ind, record in enumerate(json_data):
#         for key, start in zip(['超声', '放射'], cn_starts):
#             if key in record:
#                 item_arr = [(item['日期'], item['数据']) for item in record[key]]
#                 item_arr = sorted(item_arr, key=lambda x: x[0])
#
#                 for idx2, (_, item_data) in enumerate(item_arr):
#                     txt, match_txt, val = '', '', None
#                     for item2 in item_data:
#                         arr = item2.split(',')
#                         txt = txt + arr[-2] + arr[-1]
#
#                     if re.search(diease_regex[diease], txt):
#                         val = 1
#                         if diease in diease_suspect_regex and re.search(diease_suspect_regex[diease], txt):
#                             val = 3
#                         elif re.search(diease_regex[diease] + inner + regex_suspect, txt):
#                             val = 3
#                     else:
#                         val = 0
#
#                     total_ct = total_ct + 1
#                     if sheet.cell(rn, start + idx2).value is None and val is None:
#                         same_ct = same_ct + 1
#                     elif sheet.cell(rn, start + idx2).value is not None and val is not None:
#                         if (str(sheet.cell(rn, start + idx2).value) == '2' or str(sheet.cell(rn, start + idx2).value == '0')) and \
#                             str(val) == '0':
#                             same_ct = same_ct + 1
#                         elif str(sheet.cell(rn, start + idx2).value) == '4' and str(val) == '0':
#                             same_ct = same_ct + 1
#                         elif str(sheet.cell(rn, start + idx2).value) == str(val):
#                             same_ct = same_ct + 1
#                         else:
#                             # print(str(sheet.cell(rn, start + idx2).value), str(val))
#                             pass
#                     elif sheet.cell(rn, start + idx2).value is None and str(val) == '0':
#                         same_ct = same_ct + 1
#                     else:
#                         # print(str(sheet.cell(rn, start + idx2).value), str(val))
#                         pass
#
#         rn = rn + 1
#
#     return same_ct, total_ct

def process_records(key_file, json_file, out_path):
    """
    处理记录数据
    """
    keys = load_keys(key_file, with_head=False)
    keys = [e[0] + '_' + e[1] for e in keys]

    json_data = ''
    with open(json_file, encoding='utf-8') as f:
        json_data = json.load(f, strict=False)

    json_data = filter_records(json_data, keys)
    print(len(list(keys)), len(json_data))

    regex_dict = {'超声': '(超声)|(彩超)|(B超)', 'CT': '(CT)|(平扫)', 'MR': '(MR)|(DWI)', 'DR': '(X线)|(DR)|(钡餐)|(侧位)|(床旁片)|(平片)|(斜位)|(胸部正)|(正侧位)|(正位)'}

    isp = InspRule()
    results = []
    for item in json_data:
        text_arr_dict = {'超声': [], 'CT': [], 'MR': [], 'DR': []}
        # 超声
        if '超声' in item:
            for item_cs in item['超声']:
                for item_cs_sj in item_cs['数据']:
                    arr = item_cs_sj.split(',')
                    if arr[2] != '':
                        text = arr[2] + arr[3] if len(arr) > 3 else arr[2]
                        text_arr_dict['超声'].extend(isp.split_text(text))

        # 放射
        if '放射' in item:
            for item_fs in item['放射']:
                for item_fs_sj in item_fs['数据']:
                    arr = item_fs_sj.split(',')
                    if arr[2] != '':
                        type = arr[0].upper()
                        text = arr[2] + arr[3] if len(arr) > 3 else arr[2]
                        if 'CT' in type:
                            text_arr_dict['CT'].extend(isp.split_text(text))
                        elif 'MR' in type:
                            text_arr_dict['MR'].extend(isp.split_text(text))
                        elif 'DR' in type or 'CR' in type or 'X线' in type:
                            text_arr_dict['DR'].extend(isp.split_text(text))

        # 门诊外院
        texts = get_mzwy_texts(item)
        for text in texts:
            for key in text_arr_dict.keys():
                if re.search(regex_dict[key], text):
                    text_arr_dict[key].append(text)
                    break

        # 按Key处理
        result = {}
        for key in text_arr_dict.keys():
            if len(text_arr_dict[key]) == 0:
                text_arr_dict[key] = ['  ']

            # 影像正则处理
            key_result = isp.merge_results_arr(isp.process_arr(text_arr_dict[key]))
            for key2 in key_result.keys():
                result['%s_%s' % (key, key2)] = key_result[key2]

        results.append(result)


    write_sheet_arr_dict(results, out_path, 'Sheet1', debug=False)


if __name__ == '__main__':
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='3456', help='postfix num')
    parser.add_argument('-t', type=str, default='腹痛', help='数据类型')
    args = parser.parse_args()

    postfix = args.p
    data_type = args.t
    if not os.path.exists('data/%s' % data_type):
        print('data type: %s not exists' % data_type)
        exit()
    print("postfix: %s, data_type: %s" % (data_type, postfix))
    if not os.path.exists('data/%s/labeled_ind_%s.txt' % (data_type, postfix)):
        print('mrnos file: data/%s/labeled_ind_%s.txt not exists!' % (data_type, postfix))
        exit()

    process_records(r'data/%s/labeled_ind_%s.txt' % (data_type, postfix), r'data/%s/汇总结果_4335.json' % data_type, r'data/%s/影像学正则结果_%s.xlsx' % (data_type, postfix))




    # isp = InspRule()
    # data = load_sheet_arr_dict(r'data\腹痛\影像学正则测试数据.xlsx', 'Sheet1')
    #
    # for row in data:
    #     result = isp.process(row['文本'])
    #     for key in row.keys():
    #         if key != '文本':
    #             row[key] = result[key]
    #
    # write_sheet_arr_dict(data, r'data\腹痛\影像学正则测试数据_结果.xlsx', 'Sheet1')

    # # 加载json数据
    # json_data = ''
    # with open(r'data/汇总结果_%s.json' % postfix) as f:
    #     json_data = json.load(f, strict=False)
    # mrnos = load_mrnos(r'data/labeled_ind_1380.txt', with_head=False, sperator='	')
    # json_data = filter_json_values(json_data, mrnos)
    #
    # # # 写excel #####################
    # # workbook = load_workbook(r"data/result.xlsx")
    # wb = Workbook()
    # sheets = []
    # for idx, diease in enumerate(diease_regex.keys()):
    #     sheets.append(wb.create_sheet(diease, idx))
    #
    # _, _, num_cs, num_fs, _, _ = get_max_num(json_data)
    # for diease, sheet in zip(list(diease_regex.keys()), sheets):
    #     write_sheet_header(sheet, num_cs, num_fs)
    #     write_sheet_data(sheet, diease, json_data, num_cs, num_fs)
    #
    # # # 保存文档
    # wb.save(r'data\r_dpi_%s.xlsx' % postfix)
    # wb.close()



    # # 统计结果
    # workbook = load_workbook(r"C:\Users\Administrator\Desktop\训练集病例影像学.xlsx")
    # sheets = []
    # for idx, diease in enumerate(diease_regex.keys()):
    #     print(diease)
    #     same_ct, total_ct = compare_sheet_data(workbook[diease], diease, json_data)
    #     print(same_ct, total_ct, same_ct / total_ct)
