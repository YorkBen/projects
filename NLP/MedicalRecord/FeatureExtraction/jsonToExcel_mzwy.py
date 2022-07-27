# pip install openpyxl
from openpyxl import load_workbook
import json
import re
import os
import argparse

"""
门诊外院
"""

def load_regex():
    lines = []
    for filename in ['放射', '超声', '检验_腹痛']:
        with open(r'data\regex\%s.txt' % filename) as f:
            for l in f.readlines():
                lines.append(l.strip())

    return '((' + ')|('.join(lines) + '))'

if __name__ == '__main__':
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='2409', help='postfix num')
    parser.add_argument('-t', type=str, default='腹痛', help='数据类型')
    args = parser.parse_args()

    postfix = args.p
    data_type = args.t
    if not os.path.exists('data/%s' % data_type):
        print('data type: %s not exists' % data_type)
        exit()
    print("postfix: %s, data_type: %s" % (data_type, postfix))
    if not os.path.exists('data/%s/labeled_ind_%s.txt' % (data_type, postfix)):
        print('mrnos file: data/%s/labeled_ind_%s.txt not exists!' % (data_type, postfix))
        exit()


    # 加载json数据
    json_data = ''
    with open(r'data/%s/汇总结果_%s.json' % (data_type, postfix)) as f:
        json_data = json.load(f, strict=False)


    def get_max_num(json_data):
        """
        '日常病程', '实验室数据', '超声', '放射', '病理', '医嘱'
        """
        key_arr = ['日常病程', '实验室数据', '超声', '放射', '病理', '医嘱']
        max_len_arr = [0, 0, 0, 0, 0, 0]
        for item in json_data:
            for idx, key in enumerate(key_arr):
                if key in item and len(item[key]) > max_len_arr[idx]:
                    max_len_arr[idx] = len(item[key])
        return max_len_arr


    def get_json_value(item, key):
        if key not in item:
            return ''
        else:
            return str(item[key])


    columns = ['医保编号', '性别', '年龄', '入院时间', '主诉', '现病史', '既往史', '手术外伤史', '药物过敏史',
                    '个人史', '婚育史', '月经史', '家族史', '体格检查', '专科情况（体检）', '门诊及院外重要辅助检查',
                    '病史小结', '出院记录', '出院诊断', '首次病程']

    # 日常病程 最大33个
    # 实验室数据 最大23个
    # max_len, max_len2 = get_max_num(json_data)
    # max_len, max_len2 = 33, 23
    # for i in range(1, max_len + 1):
    #     columns.append('日常病程%s' % i)
    # for i in range(1, max_len2 + 1):
    #     columns.append('实验室数据%s' % i)
    # columns.append('出院记录')

    # # 写excel #####################
    workbook = load_workbook(r"data/result.xlsx")
    sheet, sheet2 = workbook["Sheet1"], workbook["Sheet2"]

    # sheet1
    # for ind, c in enumerate(columns):
    #     sheet.cell(1, 1+ind).value = c
    #
    # for ind, item in enumerate(json_data):
    #     rn = ind + 2
    #     cn = 0
    #     for ind_c, elem in enumerate(columns):
    #         cn = cn + 1
    #         if elem in ['医保编号', '首次病程', '出院记录']:
    #             sheet.cell(rn, cn).value = get_json_value(item, elem)
    #         elif elem == '出院诊断':
    #             if '出院记录' in item:
    #                 sheet.cell(rn, cn).value = get_json_value(item['出院记录'], elem)
    #             elif '出院诊断证明书' in item:
    #                 sheet.cell(rn, cn).value = get_json_value(item['出院诊断证明书'], elem)
    #             else:
    #                 sheet.cell(rn, cn).value = ''
    #         elif '入院记录' in item:
    #             sheet.cell(rn, cn).value = get_json_value(item['入院记录'], elem)
    #
    #     # 按照日期先后排列日常病程和实验室数据
    #     str_arr = []
    #     if '日常病程' in item:
    #         for item2 in item['日常病程']:
    #             str_arr.append((item2['DATE'], '日常病程\n' + str(item2)))
    #     for key in ['检验', '医嘱']:
    #         if key in item:
    #             for item3 in item[key]:
    #                 str_arr.append((item3['日期'], key + '\n' + str(item3)))
    #     str_arr = sorted(str_arr, key=lambda x: x[0])
    #
    #     for item23 in str_arr:
    #         cn = cn + 1
    #         sheet.cell(rn, cn).value = item23[1]




    # # sheet2
    # 表头
    sheet2.cell(1, 1).value = '医保编号'
    _, _, num_cs, num_fs, num_bl, _ = get_max_num(json_data)
    cn = 2
    # for key, num in zip(['超声', '放射', '病理'], [num_cs, num_fs, num_bl]):
    #     for i in range(1, num+1):
    #         sheet2.cell(1, cn).value = '%s_%d' % (key, i)
    #         cn = cn + 1
    sheet2.cell(1, cn).value = '门诊&外院&超声_1'
    sheet2.cell(1, cn+1).value = '门诊&外院&超声_2'
    sheet2.cell(1, cn+2).value = '门诊&外院&超声_3'
    sheet2.cell(1, cn+3).value = '门诊&外院&超声_4'
    sheet2.cell(1, cn+4).value = '门诊&外院&超声_5'


    regex = load_regex()
    cn_starts = [2, 2 + num_cs, 2 + num_cs + num_fs]
    rn = 2
    for ind, item in enumerate(json_data):
        sheet2.cell(rn, 1).value = get_json_value(item, '医保编号')

        # for key, cn in zip(['超声', '放射', '病理'], cn_starts):
        #     if key in item:
        #         str_arr = [(item_cs['日期'], str(item_cs)) for item_cs in item[key]]
        #         str_arr = sorted(str_arr, key=lambda x: x[0])
        #
        #         for item_str in str_arr:
        #             sheet2.cell(rn, cn).value = item_str[1]
        #             cn = cn + 1

        # 后面附加一列，提取外院&超声数据
        s1 = get_json_value(item['入院记录'], '门诊及院外重要辅助检查') if '入院记录' in item else ''
        s2 = get_json_value(item['入院记录']['病史小结'], '辅助检查') if '入院记录' in item and '病史小结' in item['入院记录'] else ''
        s3 = get_json_value(item['首次病程']['病例特点'], '辅助检查') if '首次病程' in item else ''
        s4 = get_json_value(item['出院记录']['入院情况'], '辅助检查') if '出院记录' in item else ''
        s5 = get_json_value(item['入院记录'], '现病史') if '入院记录' in item else ''
        # if re.search('(超声)|(彩超)|(B超)', s):
        # regex = '(医院)|(外院)|(超声)|(彩超)|(B超)'
        sheet2.cell(rn, cn).value = s1 if re.search(regex, s1) else ''
        sheet2.cell(rn, cn+1).value = s2 if re.search(regex, s2) else ''
        sheet2.cell(rn, cn+2).value = s3 if re.search(regex, s3) else ''
        sheet2.cell(rn, cn+3).value = s4 if re.search(regex, s4) else ''
        sheet2.cell(rn, cn+4).value = s5 if re.search(regex, s5) else ''

        rn = rn + 1

    # # 保存文档
    workbook.save(r"data/%s/mzwy_%s.xlsx" % (data_type, postfix))
