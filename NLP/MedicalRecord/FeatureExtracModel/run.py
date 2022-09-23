import random
import logging
import sys

sys.path.append('../../Lib/Models')

from DataLoader import DataLoader
from TextClassifier import TextClassifier
from openpyxl import load_workbook, Workbook

logging.basicConfig(level=logging.DEBUG)

def gen_train_val_data(train_lines, val_lines, text_col, label_col, feature_name, using_text_pair=True):
    """
    从原始txt数据合成训练所需格式的数据
    """
    if using_text_pair:
        train_data = [(line[text_col], feature_name, line[label_col]) for line in train_lines]
        val_data = [(line[text_col], feature_name, line[label_col]) for line in val_lines]
    else:
        train_data = [(line[text_col], line[label_col]) for line in train_lines]
        val_data = [(line[text_col], line[label_col]) for line in val_lines]
    return train_data, val_data


def cust_logging(log_file, str):
    with open(log_file, 'a+') as f:
        f.write(str + '\n')


def stat_acc_by_cls(labels, preds):
    """
    统计预测结果标签数量及正确率
    """
    result = {}
    for i1, i2 in zip(labels, preds):
        i1, i2 = int(i1), int(i2)
        if i1 not in result:
            result[i1] = {
                'count': 0,
                'correct': 0
            }
        result[i1]['count'] = result[i1]['count'] + 1
        if i1 == i2:
            result[i1]['correct'] = result[i1]['correct'] + 1

    for i in result.keys():
        result[i]['ratio'] = result[i]['correct'] / result[i]['count']

    return result

def write_stat_result(feature_name, result, log_file):
    result_arr = [(0, 0) for i in range(4)]
    total_count, total_correct = 0, 0
    for key, val in result.items():
        result_arr[key] = (val['count'], val['ratio'])
        total_count = total_count + val['count']
        total_correct = total_correct + val['correct']
    result_arr.append((total_count, total_correct/total_count))

    with open(log_file, 'a+') as f:
        s = feature_name
        for ct, ratio in result_arr:
            s = s + ('	%d	%.2f' % (ct, ratio))
        f.write('%s\n' % s)


def train1(file_path, text_field, label_field, log_file):
    dl = DataLoader()
    lines = dl.load_data_lines(file_path=file_path, num_fields=2, skip_title=False, shuffle=False)
    cols, data_lines = lines[0], lines[1:]
    feature_name = cols[1]
    random.shuffle(data_lines)
    train_lines, val_lines = dl.split_data_by_cls_num(data_lines, cls_col=1)
    train_data, val_data = gen_train_val_data(train_lines, val_lines, text_col=0, label_col=1, feature_name=feature_name)
    # 初始化模型
    model = TextClassifier(model_save_path='output/models',
                            pre_model_path="../BertModels/medical-roberta-wwm",
                            num_cls=3,
                            model_name=feature_name)

    model.load_train_val_data(train_data, val_data, label_dict={'0': 0, '1': 1, '2': 2}, batch_size=8)
    return model.train(write_result_to_file=log_file, early_stopping_num=5)


def train(train_file_path, val_file_path, text_field, label_field, num_fields, feature_name, log_file, skip_title=False, using_text_pair=True, model_save_path='output/models'):
    dl = DataLoader()
    train_lines = dl.load_data_lines(file_path=train_file_path, num_fields=num_fields, separator='	', skip_title=skip_title, shuffle=False)
    val_lines = dl.load_data_lines(file_path=val_file_path, num_fields=num_fields, separator='	', skip_title=skip_title, shuffle=False)
    train_data, val_data = gen_train_val_data(train_lines, val_lines, text_col=text_field, label_col=label_field, feature_name=feature_name, using_text_pair=using_text_pair)

    # 初始化模型
    model = TextClassifier(model_save_path=model_save_path,
                            pre_model_path="../BertModels/medical-roberta-wwm",
                            # pre_model_path="hfl/chinese-roberta-wwm-ext",
                            num_cls=3,
                            model_name=feature_name)

    model.load_train_val_data(train_data, val_data, label_dict={'0': 0, '1': 1, '2': 2}, batch_size=8)
    return model.train(write_result_to_file=log_file, early_stopping_num=5)


def predict(mdl_filepath, predict_file_path, result_file_path, log_file_path, text_field, label_field, num_fields, feature_name, skip_title=True):
    dl = DataLoader()
    lines = dl.load_data_lines(file_path=predict_file_path, num_fields=num_fields, separator='	', skip_title=True, shuffle=False)
    ids = [[line[0], line[1]] for line in lines]
    sentences = [line[text_field] for line in lines]
    labels = [line[label_field] for line in lines]
    keywords = [feature_name for line in lines]

    model = TextClassifier(model_save_path='output/models',
                            pre_model_path="../BertModels/medical-roberta-wwm",
                            num_cls=3,
                            model_name='',
                            model_file_path=mdl_filepath
                            )

    model.load_data(sentences, labels, label_dict={'0': 0, '1': 1, '2': 2}, texts_pair=keywords, batch_size=8, is_training=False)
    results = model.predict_nowrite()
    with open(result_file_path, 'w') as f:
        for (mrno, rdate), (lbl, pred) in zip(ids, results):
            f.write('%s	%s	%s	%s\n' % (mrno, rdate, lbl, pred))

    stat_result = stat_acc_by_cls([e[0] for e in results], [e[1] for e in results])
    write_stat_result(feature_name, stat_result, log_file_path)

def train_all_features(file_path, start_field, num_fields, log_file):
    dl = DataLoader()
    ## 训练TextClassifier ############################################
    # 加载所有数据
    lines = dl.load_data_lines(file_path=file_path, num_fields=num_fields, skip_title=False, shuffle=False)
    cols, data_lines = lines[0], lines[1:]
    # 训练循环
    # 生成训练数据
    for k in range(start_field, num_fields):
        feature_name = cols[k]
        cust_logging(log_file, 'Training Feature: %s' % feature_name)
        logging.debug('Training Feature: %s' % feature_name)
        random.shuffle(data_lines)
        train_lines, val_lines = dl.split_data_by_cls_num2(data_lines, k)
        # train_num = int(len(data_lines) * 0.8)
        # train_lines, val_lines = data_lines[:train_num], data_lines[train_num:]

        train_data, val_data = gen_train_val_data(train_lines, val_lines, 1, k, feature_name)
        print('train data size: %s, val data size: %s' % (len(train_data), len(val_data)))

        # 初始化模型
        model = TextClassifier(model_save_path='output/models',
                                pre_model_path="../BertModels/medical-roberta-wwm",
                                num_cls=3,
                                model_name=feature_name)

        model.load_train_val_data(train_data, val_data, label_dict={'0': 0, '1': 1, '2': 2}, batch_size=8)
        model.train(write_result_to_file=log_file, early_stopping_num=5)

    # 所有数据混合训练
    # print('mix feature training, sample nums: 0:%s, 1:%s, 2:%s' % (labels_cls_count_all[0], labels_cls_count_all[1], labels_cls_count_all[2]))
    # writeToFile(result_file, 'mix feature training, sample nums: 0:%s, 1:%s, 2:%s' % (labels_cls_count_all[0], labels_cls_count_all[1], labels_cls_count_all[2]))
    # datasets = DataToDataset(tokenizer, sentences_all, keywords_all, labels_all)
    # train(datasets, device)


def load_data_sheet(sheet):
    """
    workbook_path: excel路径
    sheet_name：表单名
    """
    txts = []
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(i, 1).value is None:
            break
        txts.append(sheet.cell(i, 2).value)

    return txts


# def predict_all_features(workbook_path):
#     dl = DataLoader()
#
#     workbook = load_workbook(workbook_path)
#     sheet = workbook[sheet_name]
#     # 第一列是医保编号，第二列是现病史，后续是特征提取
#     sentences = load_data_sheet(sheet)
#     labels = [1 for t in sentences]
#
#     ## 训练TextClassifier ############################################
#     # 加载所有数据
#     model_names = ['放射痛（右肩、肩胛和背部）_20220719145629_9016.pth', '上腹痛_20220719150741_9364.pth', '厌食_20220719152048_7950.pth', \
#                   '腹泻_20220719153533_8504.pth', '盆腔疼痛_20220719154647_9979.pth', '迅速波及全腹_20220719155751_9979.pth', \
#                   '黑便_20220719161016_7663.pth', '右下腹疼痛_20220719162130_8811.pth', '转移性右下腹疼痛_20220719163247_8831.pth', \
#                   '中腹痛_20220719164401_7971.pth', '右上腹痛_20220719165516_8483.pth', '放射痛（放射到同侧半腹或背部）_20220719172524_9549.pth', \
#                   '', '', '', \
#                   '', '', '', \
#                   ]
#
#
#     for idx, model_name in enumerate(model_names):
#         feature_name = model_name.split('_')[0]
#         # 初始化模型
#         model = TextClassifier(model_save_path='output/models/textclassify',
#                                 pre_model_path="BertModels/medical-roberta-wwm",
#                                 num_cls=3,
#                                 model_name=feature_name,
#                                 model_file_path='output/models/textclassify/%s' % model_name
#                                 )
#
#         model.load_data(sentences, labels, texts_pair=keywords, batch_size=8, is_training=False)
#         results = model.predict_nowrite(write_result_to_file='output/results/textclassify_predict_20220518.txt')
#
#         sheet.cell(1, idx + 3).value = feature_name
#         for row_idx, r in enumerate(results):
#             sheet.cell(2 + row_idx, idx + 3).value = r


if __name__ == "__main__":
    train_all_features(r'data\data_model_gen_5.txt', start_field=2, num_fields=3, log_file=r'output/textclassify_train_20220729.txt')
#     log_file = r'output/textclassify_train_20220729.txt'
#     train_all_features(r'data/data_model_2049.txt', start_field=2, num_fields=23, log_file=log_file)
    # train_all_features(r'data/data_model_2049_2.txt', start_field=2, num_fields=5)
    # train_all_features(r'data/data_model_2049_3.txt', start_field=2, num_fields=3)
    # train_all_features(r'data/data_model_2049_4.txt', start_field=2, num_fields=4)
    # predict_all_features()

    #
    #
