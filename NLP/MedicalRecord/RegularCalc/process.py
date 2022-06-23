from openpyxl import load_workbook

feature_diease = {}
feature_score = {}
diease_features = {}
with open('FeatureDieaseMap.txt', 'r') as f:
    for line in f.readlines():
        arr = line.strip().split('	')
        if len(arr) != 2:
            print(line)

        # 症状 -> 疾病
        if arr[0] not in feature_diease:
            feature_diease[arr[0]] = []
        feature_diease[arr[0]].append(arr[1])

        # 疾病 -> 症状
        if arr[1] not in diease_features:
            diease_features[arr[1]] = []
        diease_features[arr[1]].append(arr[0])

# 计算症状分数
for feature in feature_diease.keys():
    score = 1 / len(set(feature_diease[feature]))
    feature_score[feature] = score
# print(feature_score)

# 计算疾病症状
for diease in diease_features.keys():
    diease_features[diease] = list(set(diease_features[diease]))


workbook = load_workbook(r"特征库标注20220608.xlsx")
sheet1 = workbook["Sheet1"]
sheet2 = workbook["Sheet2"]

# 根据Sheet1初始化Sheet2
row_ct = 0
for rn, row in enumerate(sheet1.rows):
    if sheet1.cell(rn+1, 1).value is None:
        break
    row_ct = row_ct + 1
    for cn, cell in enumerate(row):
        # print(rn, cn, cell.value)
        if cn <= 2:
            sheet2.cell(rn+1, cn+1).value = cell.value
        else:
            break

col_ct = 0
cols = ['']
for row in sheet1.rows:
    for cell in row:
        if cell.value is not None:
            col_ct = col_ct + 1
            str = cell.value.split('；')[0]
            str = str.replace('，现病史', '').replace('，首次病程', '').replace('，主诉', '').replace('，病史小结', '')
            str = str.replace('，体格检查', '').replace('，既往史', '')
            cols.append(str)
        else:
            break
    break
print('row count: %s, col count: %s' % (row_ct, col_ct))
# print(cols)

sheet2.cell(1, 4).value = '金标准'
# 根据疾病初始化Sheet2表头
for idx, diease in enumerate(diease_features.keys()):
    sheet2.cell(1, idx+5).value = diease


def calc_score(features, feature_score, len):
    """
    匹配两个features列表，输出union部分
    """
    score = 0
    for f in features:
        score = score + feature_score[f]

    return score / len


def stat_diease_features(features, feature_score, diease_features):
    """
    根据输入的症状，症状分数字典，以及疾病所有的症状字典，统计症状对应各个疾病的分数分布
    """
    features_set = set(features)
    diease_score = {}
    for diease in diease_features.keys():
        features_ = diease_features[diease]
        intersec = list(features_set.intersection(set(features_)))
        diease_score[diease] = calc_score(intersec, feature_score, len(features_))

    return diease_score

# 金标准数据
mrno_standard = {}
with open('goldenstandard.txt') as f:
    for line in f.readlines():
        arr = line.strip().split('	')
        mrno_standard[arr[0]] = arr[1]

# 计算疾病分数
for rn in range(2, row_ct+1):
    features = []
    # 每一行一个病历
    for cn in range(4, col_ct+1):
        if sheet1.cell(rn, cn).value == 1:
            features.append(cols[cn])

    # print(features)
    diease_score = stat_diease_features(features, feature_score, diease_features)
    # print(diease_score)

    # exit()
    # 写入结果
    for idx, diease in enumerate(diease_score.keys()):
        sheet2.cell(rn, idx+5).value = diease_score[diease]


    # 写入金标准数据
    sheet2.cell(rn, 4).value = mrno_standard[sheet2.cell(rn, 1).value]




workbook.save(r"结果.xlsx")
