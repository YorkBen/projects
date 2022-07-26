import random
import logging
from Lib.DataLoader import DataLoader
from Lib.TextClassifier import TextClassifier
from openpyxl import load_workbook, Workbook

logging.basicConfig(level=logging.DEBUG)

def gen_train_val_data(train_lines, val_lines, text_col, label_col, feature_name):
    """
    从原始txt数据合成训练所需格式的数据
    """
    train_data = [(line[text_col], feature_name, line[label_col]) for line in train_lines]
    val_data = [(line[text_col], feature_name, line[label_col]) for line in val_lines]
    return train_data, val_data


def cust_logging(log_file, str):
    with open(log_file, 'a+') as f:
        f.write(str + '\n')


def train_all_features(file_path, start_field, num_fields):
    dl = DataLoader()
    ## 训练TextClassifier ############################################
    # 加载所有数据
    lines = dl.load_data_lines(file_path=file_path, num_fields=num_fields, skip_title=False, shuffle=False)
    cols, data_lines = lines[0], lines[1:]
    # 训练循环
    # 生成训练数据
    for k in range(start_field, num_fields):
        feature_name = cols[k]
        cust_logging(log_file, 'Training Feature: %s' % feature_name)
        logging.debug('Training Feature: %s' % feature_name)
        random.shuffle(data_lines)
        train_lines, val_lines = dl.split_data_by_cls_num2(data_lines, k)
        # train_num = int(len(data_lines) * 0.8)
        # train_lines, val_lines = data_lines[:train_num], data_lines[train_num:]

        train_data, val_data = gen_train_val_data(train_lines, val_lines, 1, k, feature_name)
        print('train data size: %s, val data size: %s' % (len(train_data), len(val_data)))

        # 初始化模型
        model = TextClassifier(model_save_path='output/models/textclassify',
                                pre_model_path="../../BertModels/medical-roberta-wwm",
                                num_cls=3,
                                model_name=feature_name)

        model.load_train_val_data(train_data, val_data, label_dict={'0': 0, '1': 1, '2': 2}, batch_size=8)
        model.train(write_result_to_file=log_file, early_stopping_num=5)

    # 所有数据混合训练
    # print('mix feature training, sample nums: 0:%s, 1:%s, 2:%s' % (labels_cls_count_all[0], labels_cls_count_all[1], labels_cls_count_all[2]))
    # writeToFile(result_file, 'mix feature training, sample nums: 0:%s, 1:%s, 2:%s' % (labels_cls_count_all[0], labels_cls_count_all[1], labels_cls_count_all[2]))
    # datasets = DataToDataset(tokenizer, sentences_all, keywords_all, labels_all)
    # train(datasets, device)


def load_data_sheet(sheet):
    """
    workbook_path: excel路径
    sheet_name：表单名
    """
    txts = []
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(i, 1).value is None:
            break
        txts.append(sheet.cell(i, 2).value)

    return txts


# def predict_all_features(workbook_path):
#     dl = DataLoader()
#
#     workbook = load_workbook(workbook_path)
#     sheet = workbook[sheet_name]
#     # 第一列是医保编号，第二列是现病史，后续是特征提取
#     sentences = load_data_sheet(sheet)
#     labels = [1 for t in sentences]
#
#     ## 训练TextClassifier ############################################
#     # 加载所有数据
#     model_names = ['放射痛（右肩、肩胛和背部）_20220719145629_9016.pth', '上腹痛_20220719150741_9364.pth', '厌食_20220719152048_7950.pth', \
#                   '腹泻_20220719153533_8504.pth', '盆腔疼痛_20220719154647_9979.pth', '迅速波及全腹_20220719155751_9979.pth', \
#                   '黑便_20220719161016_7663.pth', '右下腹疼痛_20220719162130_8811.pth', '转移性右下腹疼痛_20220719163247_8831.pth', \
#                   '中腹痛_20220719164401_7971.pth', '右上腹痛_20220719165516_8483.pth', '放射痛（放射到同侧半腹或背部）_20220719172524_9549.pth', \
#                   '', '', '', \
#                   '', '', '', \
#                   ]
#
#
#     for idx, model_name in enumerate(model_names):
#         feature_name = model_name.split('_')[0]
#         # 初始化模型
#         model = TextClassifier(model_save_path='output/models/textclassify',
#                                 pre_model_path="BertModels/medical-roberta-wwm",
#                                 num_cls=3,
#                                 model_name=feature_name,
#                                 model_file_path='output/models/textclassify/%s' % model_name
#                                 )
#
#         model.load_data(sentences, labels, texts_pair=keywords, batch_size=8, is_training=False)
#         results = model.predict_nowrite(write_result_to_file='output/results/textclassify_predict_20220518.txt')
#
#         sheet.cell(1, idx + 3).value = feature_name
#         for row_idx, r in enumerate(results):
#             sheet.cell(2 + row_idx, idx + 3).value = r


if __name__ == "__main__":
    log_file = r'output/textclassify_train_20220726.txt'
    train_all_features(r'data/data_model_2049_2.txt', start_field=4, num_fields=5)
    train_all_features(r'data/data_model_2049_3.txt', start_field=2, num_fields=3)
    train_all_features(r'data/data_model_2049_4.txt', start_field=2, num_fields=4)
    # predict_all_features()

    #
    #
