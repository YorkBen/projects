# pip install openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import re
import os
import sys
import argparse

sys.path.append('../Lib')

from MRRecordUtil import *
from Lib.InspRule import InspRule

# def compare_sheet_data(sheet, diease, json_data):
#     cn_starts = [2, 7]
#     rn = 2
#     same_ct, total_ct = 0, 0
#     for ind, record in enumerate(json_data):
#         for key, start in zip(['超声', '放射'], cn_starts):
#             if key in record:
#                 item_arr = [(item['日期'], item['数据']) for item in record[key]]
#                 item_arr = sorted(item_arr, key=lambda x: x[0])
#
#                 for idx2, (_, item_data) in enumerate(item_arr):
#                     txt, match_txt, val = '', '', None
#                     for item2 in item_data:
#                         arr = item2.split(',')
#                         txt = txt + arr[-2] + arr[-1]
#
#                     if re.search(diease_regex[diease], txt):
#                         val = 1
#                         if diease in diease_suspect_regex and re.search(diease_suspect_regex[diease], txt):
#                             val = 3
#                         elif re.search(diease_regex[diease] + inner + regex_suspect, txt):
#                             val = 3
#                     else:
#                         val = 0
#
#                     total_ct = total_ct + 1
#                     if sheet.cell(rn, start + idx2).value is None and val is None:
#                         same_ct = same_ct + 1
#                     elif sheet.cell(rn, start + idx2).value is not None and val is not None:
#                         if (str(sheet.cell(rn, start + idx2).value) == '2' or str(sheet.cell(rn, start + idx2).value == '0')) and \
#                             str(val) == '0':
#                             same_ct = same_ct + 1
#                         elif str(sheet.cell(rn, start + idx2).value) == '4' and str(val) == '0':
#                             same_ct = same_ct + 1
#                         elif str(sheet.cell(rn, start + idx2).value) == str(val):
#                             same_ct = same_ct + 1
#                         else:
#                             # print(str(sheet.cell(rn, start + idx2).value), str(val))
#                             pass
#                     elif sheet.cell(rn, start + idx2).value is None and str(val) == '0':
#                         same_ct = same_ct + 1
#                     else:
#                         # print(str(sheet.cell(rn, start + idx2).value), str(val))
#                         pass
#
#         rn = rn + 1
#
#     return same_ct, total_ct

def process_records(key_file, json_file, out_path, debug=0):
    """
    处理记录数据
    """
    keys, json_data = load_data(json_file, key_file)

    # regex_dict = {'超声': '(超声)|(彩超)|(B超)', 'CT': 'CT', 'MR': '(MR)|(DWI)', 'DR': '(X线)|(DR)|(钡餐)|(侧位)|(床旁片)|(平片)|(斜位)|(胸部正)|(正侧位)|(正位)'}

    isp = InspRule()
    results = []
    for item in json_data:
        text_arr_dict = {'超声': [], 'CT': [], 'MR': [], 'DR': []}

        # 超声
        if '超声' in item:
            for item_cs in item['超声']:
                for item_cs_sj in item_cs['数据']:
                    arr = item_cs_sj.split(',')
                    if arr[2] != '':
                        text = arr[2] + arr[3] if len(arr) > 3 else arr[2]
                        text_arr_dict['超声'].extend(isp.split_text(text))

        # 放射
        if '放射' in item:
            for item_fs in item['放射']:
                for item_fs_sj in item_fs['数据']:
                    arr = item_fs_sj.split(',')
                    if arr[2] != '':
                        type = arr[0].upper()
                        text = arr[2] + arr[3] if len(arr) > 3 else arr[2]
                        if 'CT' in type:
                            text_arr_dict['CT'].extend(isp.split_text(text))
                        elif 'MR' in type:
                            text_arr_dict['MR'].extend(isp.split_text(text))
                        elif 'DR' in type or 'CR' in type or 'X线' in type:
                            text_arr_dict['DR'].extend(isp.split_text(text))

        # 门诊外院
        text_arr_dict2 = get_mzwy_texts(item, r'../FeatureExtracRegex/data/regex', ['超声', 'CT', 'MR', 'DR'])
        for key, val in text_arr_dict2.items():
            text_arr_dict[key].extend(val)

        # for text in get_mzwy_texts(item):
        #     for subtext in isp.split_text_by_regex_dict(text, regex_dict):
        #         for key in text_arr_dict.keys():
        #             if re.search(regex_dict[key], subtext, re.I):
        #                 text_arr_dict[key].append(subtext)
        #                 break
        # print(text_arr_dict['DR'])

        # 按Key处理，'超声'_'急性阑尾炎'
        result = {}
        for key in text_arr_dict.keys():
            if len(text_arr_dict[key]) == 0:
                text_arr_dict[key] = ['  ']

            # 影像正则处理
            key_result = isp.merge_results_arr(isp.process_arr(text_arr_dict[key]))
            for key2 in key_result.keys():
                result['%s_%s' % (key, key2)] = key_result[key2] if debug else key_result[key2][0]

        ## for debug
        # for text in text_arr_dict['MR']:
        #     print(text)
        # print('')
        # print(result['MR_急性胆囊炎'])
        # print(result['DR_急性胰腺炎'])
        # print('')

        results.append(result)

    if debug != 'label':
        write_sheet_arr_dict(results, out_path, 'Sheet1', debug=(debug==1))


if __name__ == '__main__':
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='3456', help='postfix num')
    parser.add_argument('-t', type=str, default='腹痛', help='数据类型')
    parser.add_argument('-d', type=str, default='0', help='调试类型') # 0，1， label  label的时候是人工调试，不输出excel，只打印信息；1是excel输出调试信息，0是excel不输出调试信息
    args = parser.parse_args()

    postfix = args.p
    data_type = args.t
    if not os.path.exists('../data/%s' % data_type):
        print('data type: %s not exists' % data_type)
        exit()
    print("postfix: %s, data_type: %s" % (data_type, postfix))
    if not os.path.exists('../data/%s/labeled_ind_%s.txt' % (data_type, postfix)):
        print('mrnos file: ../data/%s/labeled_ind_%s.txt not exists!' % (data_type, postfix))
        exit()
    debug_type = args.d
    labeled_file = r'../data/%s/labeled_ind_%s.txt' % (data_type, postfix) if debug_type != 'label' else r'../data/%s/labeled_ind_%s_debug.txt' % (data_type, postfix)

    process_records(key_file=labeled_file,
                    json_file=r'../data/%s/汇总结果_%s.json' % (data_type, postfix),
                    out_path=r'data/%s/影像学正则结果_%s.xlsx' % (data_type, postfix),
                    debug=debug_type)
