from openpyxl import Workbook
import argparse
import re
import sys

sys.path.append('../Lib')

from Lib.TextStructral import TextStructral
from FileUtil import load_file, load_dict
from RegexUtil import RegexUtil
from MRRecordUtil import process_mr

utils = RegexUtil()


def get_mr_no(record):
    """
    获取医保号码
    """
    for line in record:
        if '||' in line:
            return line.split('||')[0]


def struc(postfix):
    print('开始结构化入院记录....')
    ts = TextStructral()
    records = ts.load_records('data/tmp/mr_ry_%s.txt' % postfix)
    ts.load_template('data/template/入院记录.json')
    ts.set_processor()
    results = ts.process()

    wb = Workbook()
    sheet = wb.create_sheet('Sheet1', 0)
    sheet.cell(1, 1).value = '医保编号'
    sheet.cell(1, 2).value = '入院时间'
    sheet.cell(1, 3).value = '主诉'

    rn = 1
    for i in range(len(records)):
        if i > 0 and i % 1000 == 0:
            print('%s records processed!' % (i))

        mr_no = get_mr_no(records[i])
        if mr_no is not None and '主诉' in results[i]:
            rn = rn + 1
            sheet.cell(rn, 1).value = mr_no
            sheet.cell(rn, 2).value = results[i]['入院时间']
            sheet.cell(rn, 3).value = results[i]['主诉']

    wb.save(r'data/主诉_%s.xlsx' % postfix)
    wb.close()


if __name__ == "__main__":
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='all', help='postfix num')
    args = parser.parse_args()

    postfix = args.p

    process_mr(file_path=r"data/medical_record.csv", with_head=True,
                type_regex_and_outpath=[
                    ('入.*院记录', r"data/tmp/mr_ry_%s.txt" % postfix),
                    # ('出院记录', r"data/tmp/mr_cy_%s.txt" % postfix),
                    # ('首次病程', r"data/tmp/mr_sc_%s.txt" % postfix),
                    # ('日常病程', r"data/tmp/mr_rc_%s.txt" % postfix)
                ])

    struc(postfix)
