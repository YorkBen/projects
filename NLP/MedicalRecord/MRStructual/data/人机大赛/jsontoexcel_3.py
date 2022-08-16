# pip install openpyxl
import sys
import os
import re
import json
import argparse
sys.path.append('../..')
from openpyxl import load_workbook

from Lib.MRRecordUtil import process_mr, load_mrno


# 加载json数据
json_data = ''
with open(r'汇总结果_231.json') as f:
    json_data = json.load(f, strict=False)


def get_json_value(item, key):
    if key not in item:
        return ''
    else:
        return str(item[key])

def check_cell_empty(value):
    if value is None:
        return True
    elif value == 'None':
        return True
    elif value == '':
        return True
    return False


columns = [
    '医保编号',
    ['性别', '年龄', '入院时间'],
    '姓名'
]

if __name__ == '__main__':
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='2409', help='postfix num')
    args = parser.parse_args()

    postfix = args.p
    print("postfix: %s" % (postfix))
    if not os.path.exists('labeled_ind_%s.txt' % postfix):
        print('mrnos file: labeled_ind_%s.txt not exists!' % postfix)
        exit()

    # 加载mrnos
    mr_nos = load_mrno('labeled_ind_%s.txt' % postfix, with_head=False)

    # 写excel
    workbook = load_workbook(r"../result.xlsx")
    sheet = workbook["Sheet1"]

    # 写表头
    ind = 1
    for col in columns:
        if isinstance(col, list):
            for col2 in col:
                sheet.cell(1, ind).value = col2
                ind = ind + 1
        else:
            sheet.cell(1, ind).value = col
            ind = ind + 1

    ## 写内容 #############
    for ind, item in enumerate(json_data):
        rn = ind + 2
        # 医保编号
        sheet.cell(rn, 1).value = get_json_value(item, '医保编号')

        # 入院记录
        cn = 2
        for col in columns[1]:
            sheet.cell(rn, cn).value = get_json_value(item['入院记录'], col)
            cn = cn + 1

        # 姓名
        sheet.cell(rn, cn).value = get_json_value(item['入院记录']['病史小结'], '患者姓名')

    workbook.save(r"r_%s.xlsx" % postfix)
