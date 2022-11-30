# pip install openpyxl
from openpyxl import load_workbook
import json
import re
import os
import argparse
import sys

sys.path.append('../Lib')

from MRRecordUtil import *


if __name__ == '__main__':
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='2409', help='postfix num')
    parser.add_argument('-t', type=str, default='腹痛', help='数据类型')
    args = parser.parse_args()

    postfix = args.p
    data_type = args.t
    if not os.path.exists('data/%s' % data_type):
        print('data type: %s not exists' % data_type)
        exit()
    # print("postfix: %s, data_type: %s" % (data_type, postfix))
    # if not os.path.exists('data/%s/labeled_ind_%s.txt' % (data_type, postfix)):
    #     print('mrnos file: data/%s/labeled_ind_%s.txt not exists!' % (data_type, postfix))
    #     exit()


    # 加载json数据
    json_data = ''
    with open(r'../data/%s/汇总结果_%s.json' % (data_type, postfix)) as f:
        json_data = json.load(f, strict=False)


    def get_max_num(json_data):
        """
        '日常病程', '实验室数据', '超声', '放射', '病理', '医嘱'
        """
        key_arr = ['日常病程', '实验室数据', '超声', '放射', '病理', '医嘱']
        max_len_arr = [0, 0, 0, 0, 0, 0]
        for item in json_data:
            for idx, key in enumerate(key_arr):
                if key in item and len(item[key]) > max_len_arr[idx]:
                    max_len_arr[idx] = len(item[key])
        return max_len_arr


    columns = [ ['入院记录', '一般信息', '住院号'],
                ['入院记录', '一般信息', '床号'],
                ['入院记录', '一般信息', '姓名'],
                ['入院记录', '一般信息', '性别'],
                ['入院记录', '一般信息', '年龄'],
                ['入院记录', '一般信息', '入院时间'],
                ['入院记录', '主诉'],
                ['入院记录', '现病史'],
                ['入院记录', '既往史'],
                ['入院记录', '手术外伤史'],
                ['入院记录', '药物过敏史'],
                ['入院记录', '个人史'],
                ['入院记录', '婚育史'],
                ['入院记录', '月经史'],
                ['入院记录', '家族史'],
                ['入院记录', '体格检查'],
                ['入院记录', '专科情况'],
                ['入院记录', '门诊及院外重要辅助检查'],
                ['入院记录', '病史小结'],
                ['出院记录'],
                ['出院记录', '出院诊断'],
                ['首次病程记录'],
                ['日常病程记录'],
                ['检验'],
                ['放射'],
                # ['超声'],
                # ['其他检查']
                ['病理'],
                ['医嘱']
            ]

    # 日常病程 最大33个
    # 实验室数据 最大23个
    # max_len, max_len2 = get_max_num(json_data)
    # max_len, max_len2 = 33, 23
    # for i in range(1, max_len + 1):
    #     columns.append('日常病程%s' % i)
    # for i in range(1, max_len2 + 1):
    #     columns.append('实验室数据%s' % i)
    # columns.append('出院记录')

    # # 写excel #####################
    workbook = load_workbook(r"data/result.xlsx")
    sheet, sheet2 = workbook["Sheet1"], workbook["Sheet2"]

    # sheet1
    for ind, c in enumerate(columns):
        sheet.cell(1, 1+ind).value = c[-1]

    for ind, item in enumerate(json_data):
        rn = ind + 2
        cn = 0
        for ind_c, elem in enumerate(columns):
            cn = cn + 1
            val = get_json_value(item, elem)
            if elem[-1] == '入院时间' and val == '':
                val = get_json_value(item, ['出院记录', '一般信息', '入院时间'])
            sheet.cell(rn, cn).value = val

        # # 按照日期先后排列日常病程和实验室数据
        # str_arr = []
        # if '日常病程' in item:
        #     for item2 in item['日常病程']:
        #         str_arr.append((item2['DATE'], '日常病程\n' + str(item2)))
        # for key in ['检验', '医嘱']:
        #     if key in item:
        #         for item3 in item[key]:
        #             str_arr.append((item3['日期'], key + '\n' + str(item3)))
        # str_arr = sorted(str_arr, key=lambda x: x[0])
        #
        # for item23 in str_arr:
        #     cn = cn + 1
        #     sheet.cell(rn, cn).value = item23[1]



    # # # sheet2
    # # 表头
    # sheet2.cell(1, 1).value = '医保编号'
    # _, _, num_cs, num_fs, num_bl, _ = get_max_num(json_data)
    # cn = 2
    # for key, num in zip(['超声', '放射', '病理'], [num_cs, num_fs, num_bl]):
    #     for i in range(1, num+1):
    #         sheet2.cell(1, cn).value = '%s_%d' % (key, i)
    #         cn = cn + 1
    #
    #
    # cn_starts = [2, 2 + num_cs, 2 + num_cs + num_fs]
    # rn = 2
    # for ind, item in enumerate(json_data):
    #     sheet2.cell(rn, 1).value = get_json_value(item, '医保编号')
    #
    #     for key, cn in zip(['超声', '放射', '病理'], cn_starts):
    #         if key in item:
    #             str_arr = [(item_cs['日期'], str(item_cs)) for item_cs in item[key]]
    #             str_arr = sorted(str_arr, key=lambda x: x[0])
    #
    #             for item_str in str_arr:
    #                 sheet2.cell(rn, cn).value = item_str[1]
    #                 cn = cn + 1
    #
    #     rn = rn + 1

    # # 保存文档
    workbook.save(r"data/%s/病历数据/原始病历数据_%s.xlsx" % (data_type, postfix))
