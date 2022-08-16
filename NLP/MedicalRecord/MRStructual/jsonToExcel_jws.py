# pip install openpyxl
from openpyxl import load_workbook
import json
import re
import os
import argparse

def assemble_jws_str(record):
    """
    拼接既往史字符串
    """
    result_str = record['既往史'] + ' ' + record['平时健康状况'] + ' ' + record['预防接种史'].replace('无', '') + ' ' + record['预防接种药品'].replace('无', '')
    for key in record['疾病史'].keys():
        if record['疾病史'][key].replace('无', '').strip() != '':
            result_str = result_str + '。' + record['疾病史'][key]

    if result_str == '':
        result_str = '一般'

    result_str = result_str.replace(',', '，').replace('。。', '。')

    return result_str

def assemble_sss_str(record):
    """
    拼接手术史字符串
    """
    result_str = ''
    for key in record.keys():
        if record[key].replace('有', '').strip() != '':
            result_str = result_str + record[key] + '。'
    result_str = result_str.replace(',', '，').replace('。。', '。')

    return result_str


if __name__ == '__main__':
    # 参数
    parser = argparse.ArgumentParser(description='Select Data With MR_NOs')
    parser.add_argument('-p', type=str, default='2409', help='postfix num')
    parser.add_argument('-t', type=str, default='腹痛', help='数据类型')
    args = parser.parse_args()

    postfix = args.p
    data_type = args.t
    if not os.path.exists('data/%s' % data_type):
        print('data type: %s not exists' % data_type)
        exit()
    print("postfix: %s, data_type: %s" % (data_type, postfix))
    if not os.path.exists('data/%s/labeled_ind_%s.txt' % (data_type, postfix)):
        print('mrnos file: data/%s/labeled_ind_%s.txt not exists!' % (data_type, postfix))
        exit()


    # 加载json数据
    json_data = ''
    with open(r'data/%s/汇总结果_%s.json' % (data_type, postfix)) as f:
        json_data = json.load(f, strict=False)



    def get_json_value(item, key):
        if key not in item:
            return ''
        else:
            return str(item[key])


    columns = ['医保编号', '既往史', '手术外伤史', '现病史+既往史+手术外伤史']


    # # 写excel #####################
    workbook = load_workbook(r"data/result.xlsx")
    sheet= workbook["Sheet1"]

    # sheet1
    for ind, c in enumerate(columns):
        sheet.cell(1, 1+ind).value = c

    for ind, item in enumerate(json_data):
        rn = ind + 2
        sheet.cell(rn, 1).value = get_json_value(item, '医保编号')
        # 既往史
        str_jws = assemble_jws_str(item['入院记录']['既往史'])
        sheet.cell(rn, 2).value = str_jws

        # 手术外伤史
        str_sss = assemble_sss_str(item['入院记录']['手术外伤史'])
        sheet.cell(rn, 3).value = str_sss

        # 现病史+既往史+手术外伤史
        str_xbs = item['入院记录']['现病史']
        sheet.cell(rn, 4).value = str_xbs

    # # 保存文档
    workbook.save(r"data/%s/jws_%s.xlsx" % (data_type, postfix))
